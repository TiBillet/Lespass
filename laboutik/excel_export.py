"""
Export Excel des rapports de cloture.
Genere un fichier .xlsx en bytes a partir d'une ClotureCaisse.
/ Excel export for closure reports.
Generates an .xlsx file as bytes from a ClotureCaisse.

LOCALISATION : laboutik/excel_export.py

Reutilise la meme logique que Administration/admin/laboutik.py:exporter_excel
mais retourne des bytes au lieu d'un HttpResponse.
/ Reuses the same logic as Administration/admin/laboutik.py:exporter_excel
but returns bytes instead of HttpResponse.
"""
import io

from laboutik.models import ClotureCaisse
from laboutik.reports import RapportComptableService


def _euros(centimes):
    """Convertit des centimes (int) en euros (float) pour l'affichage Excel.
    / Converts cents (int) to euros (float) for Excel display.
    """
    if centimes is None:
        return 0.0
    return centimes / 100


def generer_excel_cloture(cloture: ClotureCaisse) -> bytes:
    """
    Genere le rapport de cloture au format Excel (.xlsx).
    Retourne les bytes du fichier.
    / Generates the closure report as Excel (.xlsx).
    Returns the file bytes.

    LOCALISATION : laboutik/excel_export.py
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils.translation import gettext as _

    # Recalculer le rapport a la volee (meme logique que l'admin)
    # / Recalculate the report on-the-fly (same logic as admin)
    service = RapportComptableService(
        point_de_vente=cloture.point_de_vente,
        datetime_debut=cloture.datetime_ouverture,
        datetime_fin=cloture.datetime_cloture,
    )
    rapport = service.generer_rapport_complet()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport"

    # Styles
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    current_row = [1]
    max_cols_used = [1]
    e = _euros

    # Writer adapter (meme interface que l'admin)
    # / Writer adapter (same interface as admin)
    def append_title(titre):
        row = current_row[0]
        cell = ws.cell(row=row, column=1, value=titre)
        cell.font = section_font
        cell.fill = section_fill
        for col in range(2, 11):
            ws.cell(row=row, column=col).fill = section_fill
        current_row[0] += 1

    def append_header(cols):
        row = current_row[0]
        for idx, col_val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=idx, value=col_val)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        max_cols_used[0] = max(max_cols_used[0], len(cols))
        current_row[0] += 1

    def append_row(cols):
        row = current_row[0]
        for idx, col_val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=idx, value=col_val)
            cell.border = thin_border
            if isinstance(col_val, (int, float)):
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'
        max_cols_used[0] = max(max_cols_used[0], len(cols))
        current_row[0] += 1

    def append_blank():
        current_row[0] += 1

    # --- En-tete / Header ---
    append_title(str(_("Closure report")))
    append_row([str(_("Level")), cloture.get_niveau_display()])
    append_row([str(_("Number")), f"#{cloture.numero_sequentiel}"])
    append_row([str(_("Point of sale")), cloture.point_de_vente.name if cloture.point_de_vente else "—"])
    append_row([str(_("Period")), f"{cloture.datetime_ouverture} → {cloture.datetime_cloture}"])
    append_blank()

    # --- Section 1 : Totaux par moyen de paiement ---
    section = rapport.get("totaux_par_moyen", {})
    if section:
        append_title(str(_("Totals by payment method")))
        append_header([str(_("Payment method")), str(_("Amount"))])
        append_row([str(_("Cash")), e(section.get("especes", 0))])
        append_row([str(_("Credit card")), e(section.get("carte_bancaire", 0))])
        append_row([str(_("Cashless")), e(section.get("cashless", 0))])
        append_row([str(_("Check")), e(section.get("cheque", 0))])
        append_row([str(_("Total")), e(section.get("total", 0))])
        append_blank()

    # --- Section 4 : Solde caisse ---
    section = rapport.get("solde_caisse", {})
    if section:
        append_title(str(_("Cash register balance")))
        append_header([str(_("Item")), str(_("Amount"))])
        append_row([str(_("Opening float")), e(section.get("fond_de_caisse", 0))])
        append_row([str(_("Cash income")), e(section.get("entrees_especes", 0))])
        if section.get("sorties_especes"):
            append_row([str(_("Cash withdrawals")), -e(section.get("sorties_especes", 0))])
        append_row([str(_("Balance")), e(section.get("solde", 0))])
        append_blank()

    # --- Section 3 : TVA ---
    section = rapport.get("tva", {})
    if section:
        append_title(str(_("VAT breakdown")))
        append_header([str(_("Rate")), str(_("HT")), str(_("VAT")), str(_("TTC"))])
        for taux, data in section.items():
            append_row([taux, e(data.get("total_ht", 0)), e(data.get("total_tva", 0)), e(data.get("total_ttc", 0))])
        append_blank()

    # Auto-largeur des colonnes / Auto-width columns
    for col_idx in range(1, max_cols_used[0] + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

    # Retourner les bytes / Return the bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
