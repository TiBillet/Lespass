"""
Export Excel (.xlsx) d'une cloture comptable.
/ Excel (.xlsx) export of an accounting closure.

LOCALISATION : comptabilite/excel_export.py

Utilise openpyxl. Une seule feuille 'Rapport' avec sections empilees.
/ Single 'Rapport' sheet with stacked sections.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Styles reutilises (definis au niveau module : crees une seule fois).
# / Reusable styles defined at module level (created once).
_FONT_TITRE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
_FILL_TITRE = PatternFill("solid", fgColor="333333")
_FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FILL_SECTION = PatternFill("solid", fgColor="333333")
_FONT_HEADER = Font(name="Calibri", size=10, bold=True)
_FILL_HEADER = PatternFill("solid", fgColor="F0F0F0")
_ALIGN_RIGHT = Alignment(horizontal="right")
_BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _euros(centimes):
    if centimes is None:
        return 0.0
    return round(centimes / 100, 2)


def _ecrire_section_header(ws, row, titre, span=4):
    """Ecrit un titre de section sur 1 ligne fusionnee."""
    ws.cell(row=row, column=1, value=titre).font = _FONT_SECTION
    ws.cell(row=row, column=1).fill = _FILL_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def _ecrire_ligne_header(ws, row, colonnes):
    """Ecrit une ligne d'en-tetes de colonnes."""
    for col_idx, val in enumerate(colonnes, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.border = _BORDER_THIN
    return row + 1


def _ecrire_ligne_donnees(ws, row, valeurs, montants_indices=None):
    """Ecrit une ligne de donnees. montants_indices : colonnes alignees a droite + format nombre."""
    montants_indices = montants_indices or []
    for col_idx, val in enumerate(valeurs, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = _BORDER_THIN
        if col_idx in montants_indices:
            cell.alignment = _ALIGN_RIGHT
            cell.number_format = "#,##0.00"
    return row + 1


def generer_excel_cloture(cloture) -> tuple:
    """
    Retourne (bytes, filename, content_type) pour l'export Excel.
    / Returns (bytes, filename, content_type) for the Excel export.
    """
    rapport = cloture.rapport_json or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    row = 1

    # Titre + meta
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=f"Clôture #{cloture.numero_sequentiel} - {cloture.get_niveau_display()}")
    cell.font = _FONT_TITRE
    cell.fill = _FILL_TITRE
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=f"Début : {cloture.datetime_debut:%Y-%m-%d %H:%M} - Fin : {cloture.datetime_fin:%Y-%m-%d %H:%M}").font = Font(italic=True)
    row += 2

    # Totaux par moyen
    row = _ecrire_section_header(ws, row, "Totaux par moyen de paiement")
    row = _ecrire_ligne_header(ws, row, ["Code", "Libellé", "Total (EUR)", "Nombre"])
    for code, item in (rapport.get("totaux_par_moyen") or {}).items():
        if code in ("total", "currency_code"):
            continue
        if isinstance(item, dict):
            row = _ecrire_ligne_donnees(
                ws, row,
                [code, item.get("label", ""), _euros(item.get("total")), item.get("nb", 0)],
                montants_indices=[3],
            )
    row += 1

    # TVA
    row = _ecrire_section_header(ws, row, "Ventilation TVA")
    row = _ecrire_ligne_header(ws, row, ["Taux %", "HT (EUR)", "TVA (EUR)", "TTC (EUR)"])
    for taux, item in (rapport.get("tva") or {}).items():
        if isinstance(item, dict):
            row = _ecrire_ligne_donnees(
                ws, row,
                [item.get("taux", taux),
                 _euros(item.get("total_ht")),
                 _euros(item.get("total_tva")),
                 _euros(item.get("total_ttc"))],
                montants_indices=[2, 3, 4],
            )
    row += 1

    # Detail des ventes par categorie (depuis detail_ventes)
    # / Sales detail per category (from detail_ventes)
    row = _ecrire_section_header(ws, row, "Détail des ventes par catégorie")
    row = _ecrire_ligne_header(ws, row, ["Catégorie", "Produit", "Quantité", "HT", "TVA", "TTC"])
    for cat in (rapport.get("detail_ventes") or {}).values():
        if not isinstance(cat, dict):
            continue
        for article in cat.get("articles", []):
            row = _ecrire_ligne_donnees(
                ws, row,
                [
                    cat.get("nom_categorie", ""),
                    article.get("nom_produit", ""),
                    float(article.get("qty_total", 0)),
                    _euros(article.get("total_ht")),
                    _euros(article.get("total_tva")),
                    _euros(article.get("total_ttc")),
                ],
                montants_indices=[3, 4, 5, 6],
            )
    row += 1

    # Remboursements
    row = _ecrire_section_header(ws, row, "Remboursements et avoirs")
    row = _ecrire_ligne_header(ws, row, ["Type", "Total (EUR)", "Nombre", ""])
    rb = rapport.get("remboursements") or {}
    cn = rb.get("credit_notes", {})
    rf = rb.get("refunded", {})
    row = _ecrire_ligne_donnees(ws, row, ["Avoirs", _euros(cn.get("total")), cn.get("nb", 0), ""], montants_indices=[2])
    row = _ecrire_ligne_donnees(ws, row, ["Remboursements", _euros(rf.get("total")), rf.get("nb", 0), ""], montants_indices=[2])

    # Largeurs de colonnes fixes (5 colonnes pour la section produit/tarif)
    # / Fixed column widths (5 cols for product/tariff section)
    for col_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)

    filename = f"cloture-{cloture.numero_sequentiel}-{cloture.datetime_fin:%Y%m%d}.xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return buffer.getvalue(), filename, content_type
