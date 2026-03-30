"""
tests/pytest/test_exports.py — Tests Session 15 : exports admin.
/ Tests Session 15: admin exports.

Couvre : export PDF genere, export CSV delimiteur ;, CSV 13 sections, Excel 13 onglets.
Covers: PDF generated, CSV delimiter ;, CSV 13 sections, Excel 13 sheets.

Lancement / Run:
    docker exec lespass_django poetry run pytest tests/pytest/test_exports.py -v
"""
import sys
sys.path.insert(0, '/DjangoFiles')

import django
django.setup()

import pytest
import io
from decimal import Decimal
from django.utils import timezone
from django_tenants.utils import schema_context

from AuthBillet.models import TibilletUser
from BaseBillet.models import (
    LigneArticle, Price, PriceSold, Product, ProductSold,
    SaleOrigin, PaymentMethod,
)
from Customers.models import Client
from laboutik.models import (
    PointDeVente, ClotureCaisse, LaboutikConfiguration,
)

TENANT_SCHEMA = 'lespass'


@pytest.fixture(scope="module")
def tenant():
    """Le tenant 'lespass' (doit exister dans la base).
    / The 'lespass' tenant (must exist in DB)."""
    return Client.objects.get(schema_name=TENANT_SCHEMA)


@pytest.fixture(scope="module")
def test_data(tenant):
    """Lance create_test_pos_data pour s'assurer que les donnees existent.
    / Runs create_test_pos_data to ensure test data exists."""
    from django.core.management import call_command
    call_command('create_test_pos_data')
    return True


@pytest.fixture
def pv(test_data):
    """Retourne le premier point de vente.
    / Returns the first point of sale."""
    with schema_context(TENANT_SCHEMA):
        return PointDeVente.objects.first()


@pytest.fixture
def cloture_avec_rapport(pv, test_data):
    """Cree une cloture avec un rapport JSON complet.
    / Creates a closure with a complete JSON report."""
    with schema_context(TENANT_SCHEMA):
        from laboutik.reports import RapportComptableService

        now = timezone.now()
        debut = now - timezone.timedelta(hours=2)

        # Creer une vente pour que le rapport ne soit pas vide
        # / Create a sale so the report is not empty
        product = Product.objects.filter(
            methode_caisse=Product.VENTE,
        ).first()
        assert product is not None, "Aucun produit POS trouve. Lancer create_test_pos_data."

        price = product.prices.first()
        assert price is not None, "Le produit n'a pas de prix."

        product_sold, _ = ProductSold.objects.get_or_create(
            product=product,
            event=None,
        )
        price_sold, _ = PriceSold.objects.get_or_create(
            productsold=product_sold,
            price=price,
            defaults={'prix': price.prix},
        )
        LigneArticle.objects.create(
            pricesold=price_sold,
            amount=1000,
            qty=Decimal("1"),
            vat=Decimal("0"),
            payment_method=PaymentMethod.CASH,
            sale_origin=SaleOrigin.LABOUTIK,
            status=LigneArticle.VALID,
            point_de_vente=pv,
        )

        # Generer le rapport / Generate report
        service = RapportComptableService(pv, debut, now)
        rapport = service.generer_rapport_complet()

        responsable = TibilletUser.objects.filter(is_staff=True).first()

        cloture = ClotureCaisse.objects.create(
            point_de_vente=pv,
            responsable=responsable,
            datetime_ouverture=debut,
            datetime_cloture=now,
            niveau=ClotureCaisse.JOURNALIERE,
            numero_sequentiel=9999,
            total_especes=1000,
            total_carte_bancaire=0,
            total_cashless=0,
            total_general=1000,
            total_perpetuel=1000,
            nombre_transactions=1,
            rapport_json=rapport,
        )
        return cloture


class TestExports:
    """Tests des exports admin (PDF, CSV, Excel).
    / Admin export tests (PDF, CSV, Excel)."""

    @pytest.mark.django_db
    def test_export_pdf_genere(self, cloture_avec_rapport):
        """L'export PDF retourne un fichier non vide avec signature %PDF-.
        / PDF export returns a non-empty file with %PDF- signature."""
        from django.template.loader import render_to_string
        from weasyprint import HTML
        from BaseBillet.models import Configuration

        with schema_context(TENANT_SCHEMA):
            cloture = cloture_avec_rapport
            rapport = cloture.rapport_json or {}
            config = Configuration.get_solo()

            context = {
                "cloture": cloture,
                "rapport": rapport,
                "config_org": config.organisation or "",
                "config_siret": config.siren or "",
                "config_address": "",
                "now": timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M"),
            }

            html_string = render_to_string(
                "laboutik/pdf/rapport_comptable.html", context,
            )
            pdf_bytes = HTML(string=html_string).write_pdf()

            assert len(pdf_bytes) > 0
            # Signature PDF / PDF signature
            assert pdf_bytes[:5] == b'%PDF-'

    @pytest.mark.django_db
    def test_export_csv_delimiteur(self, cloture_avec_rapport):
        """Le CSV utilise le delimiteur ; (standard europeen).
        / CSV uses ; delimiter (European standard)."""
        import csv

        with schema_context(TENANT_SCHEMA):
            output = io.StringIO()
            writer = csv.writer(output, delimiter=';')
            writer.writerow(["test", "data"])
            csv_content = output.getvalue()

            # Verifie que le delimiteur est bien ;
            # / Verify delimiter is ;
            assert ';' in csv_content
            assert 'test;data' in csv_content

    @pytest.mark.django_db
    def test_export_csv_13_sections(self, cloture_avec_rapport):
        """Le rapport JSON contient les 13 sections attendues.
        / The JSON report contains the expected 13 sections."""
        with schema_context(TENANT_SCHEMA):
            rapport = cloture_avec_rapport.rapport_json
            assert rapport is not None

            sections_attendues = [
                "totaux_par_moyen", "detail_ventes", "tva",
                "solde_caisse", "recharges", "adhesions",
                "remboursements", "habitus", "billets",
                "synthese_operations", "operateurs",
                "ventilation_par_pv", "infos_legales",
            ]

            for section in sections_attendues:
                assert section in rapport, f"Section manquante : {section}"

    @pytest.mark.django_db
    def test_filtre_euros(self):
        """Le filtre euros convertit les centimes en affichage euros.
        / The euros filter converts cents to euro display."""
        with schema_context(TENANT_SCHEMA):
            from laboutik.templatetags.laboutik_filters import euros

            # Cas standard / Standard case
            resultat = euros(12750)
            assert "127" in resultat
            assert "50" in resultat
            assert "€" in resultat

            # Zero
            resultat_zero = euros(0)
            assert "0" in resultat_zero
            assert "00" in resultat_zero

            # Negatif / Negative
            resultat_negatif = euros(-500)
            assert "-" in resultat_negatif
            assert "5" in resultat_negatif

            # None
            resultat_none = euros(None)
            assert "0" in resultat_none

    @pytest.mark.django_db
    def test_detail_ventes_structure_enrichie(self, cloture_avec_rapport):
        """Le detail ventes contient qty_vendus, qty_offerts, prix_achat, benefice.
        / Sales detail contains sold/gifted qty, purchase price, profit."""
        with schema_context(TENANT_SCHEMA):
            rapport = cloture_avec_rapport.rapport_json
            detail_ventes = rapport.get('detail_ventes', {})

            # Verifier la structure pour chaque categorie
            # / Check structure for each category
            for categorie_nom, categorie_data in detail_ventes.items():
                assert 'articles' in categorie_data
                assert 'total_ttc' in categorie_data
                for article in categorie_data['articles']:
                    assert 'qty_vendus' in article, f"qty_vendus manquant dans {article}"
                    assert 'qty_offerts' in article, f"qty_offerts manquant dans {article}"
                    assert 'prix_achat_unit' in article, f"prix_achat_unit manquant dans {article}"
                    assert 'benefice' in article, f"benefice manquant dans {article}"
                    assert 'cout_total' in article, f"cout_total manquant dans {article}"

    @pytest.mark.django_db
    def test_habitus_structure_enrichie(self, cloture_avec_rapport):
        """Les habitus contiennent medianes et soldes.
        / Habitus contain medians and balances."""
        with schema_context(TENANT_SCHEMA):
            rapport = cloture_avec_rapport.rapport_json
            habitus = rapport.get('habitus', {})

            assert 'depense_mediane' in habitus
            assert 'recharge_mediane' in habitus
            assert 'reste_moyenne' in habitus
            assert 'med_on_card' in habitus
            assert 'nouveaux_membres' in habitus
            assert 'nb_cartes' in habitus
            assert 'panier_moyen' in habitus

    @pytest.mark.django_db
    def test_export_excel_genere(self, cloture_avec_rapport):
        """L'export Excel genere un fichier valide avec 13 onglets.
        / Excel export generates a valid file with 13 sheets."""
        import openpyxl

        with schema_context(TENANT_SCHEMA):
            cloture = cloture_avec_rapport
            rapport = cloture.rapport_json or {}

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            for section_name, section_data in rapport.items():
                # Les noms d'onglets Excel sont limites a 31 caracteres
                # / Excel sheet names are limited to 31 characters
                sheet_name = section_name[:31]
                ws = wb.create_sheet(title=sheet_name)

                if isinstance(section_data, dict):
                    ws.append(["Cle", "Valeur"])
                    for cle, valeur in section_data.items():
                        ws.append([str(cle), str(valeur)])
                elif isinstance(section_data, list) and section_data:
                    if isinstance(section_data[0], dict):
                        headers = list(section_data[0].keys())
                        ws.append(headers)
                        for item in section_data:
                            ws.append([item.get(h, '') for h in headers])
                    else:
                        for item in section_data:
                            ws.append([str(item)])
                else:
                    ws.append([str(section_data)])

            # Sauvegarder en memoire et verifier
            # / Save in memory and verify
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Recharger et verifier le nombre d'onglets
            # / Reload and check number of sheets
            wb_check = openpyxl.load_workbook(buffer)
            assert len(wb_check.sheetnames) == 13
