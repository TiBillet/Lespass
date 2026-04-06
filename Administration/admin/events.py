import logging
from datetime import timedelta
from decimal import Decimal

from django.contrib import admin, messages
from django.db import models, IntegrityError, connection
from django.db.models import Count, Q, Prefetch
from django.forms import ModelForm
from django.http import HttpRequest
from django.shortcuts import redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from import_export.admin import ImportExportModelAdmin
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget
from unfold.admin import ModelAdmin, TabularInline
from unfold.contrib.filters.admin import RangeDateTimeFilter
from unfold.contrib.forms.widgets import WysiwygWidget
from unfold.contrib.import_export.forms import ExportForm, ImportForm
from unfold.decorators import display, action
from unfold.sections import TableSection

from rest_framework import serializers as drf_serializers

from Administration.admin.site import staff_admin_site, sanitize_textfields
from ApiBillet.permissions import TenantAdminPermissionWithRequest
from ApiBillet.serializers import get_or_create_price_sold
from BaseBillet.models import (
    Configuration, Event, Ticket, Reservation, PostalAddress, LigneArticle
)

logger = logging.getLogger(__name__)


# Serializer de validation pour l'export par periode (au niveau module, pas dans EventAdmin).
# / Validation serializer for period export (module-level, not inside EventAdmin).
class ExportPeriodeSerializer(drf_serializers.Serializer):
    date_debut = drf_serializers.DateField(
        required=True,
        help_text=_("Start date (event date)"),
    )
    date_fin = drf_serializers.DateField(
        required=True,
        help_text=_("End date (event date)"),
    )
    format_export = drf_serializers.ChoiceField(
        choices=['csv', 'pdf', 'excel'],
        required=True,
    )

    def validate(self, data):
        if data['date_debut'] > data['date_fin']:
            raise drf_serializers.ValidationError(
                _("Start date must be before end date.")
            )
        return data


class EventChildrenInline(TabularInline):
    model = Event
    fk_name = 'parent'
    verbose_name = _("Volunteering")  # Pour l'instant, les enfants sont forcément des Actions.
    hide_title = True
    fields = (
        'name',
        'datetime',
        'jauge_max',
        'valid_tickets_count',
    )

    # ordering_field = "weight"
    # max_num = 1
    extra = 0
    show_change_link = True
    tab = True

    readonly_fields = (
        'valid_tickets_count',
    )

    # Surcharger la méthode pour désactiver la suppression
    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


class EventForm(ModelForm):
    class Meta:
        model = Event
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['products'].widget.can_change_related = False
        self.fields['products'].widget.can_add_related = False
        self.fields['products'].help_text = _("Leave empty to avoid reservations.")
        self.fields['short_description'].help_text = _("Used for social network descriptions.")

        try:
            # On mets la valeur de la jauge réglée dans la config par default
            config = Configuration.get_solo()
            self.fields['jauge_max'].initial = config.jauge_max
        except Exception as e:
            logger.error(f"set gauge max error : {e}")
            pass


class EventPricesSummaryTable(TableSection):
    verbose_name = _("Résumé par tarif")
    height = 240
    related_name = "pricesold_for_sections"  # Event property returning Ticket queryset with annotations
    fields = ["price_name", "qty_reserved", "total_euros"]

    def price_name(self, instance: Ticket):
        # Prefer annotated name to avoid extra queries
        name = getattr(instance, "section_price_name", None)
        if name:
            return name
        try:
            return instance.pricesold.price.name if instance.pricesold and instance.pricesold.price else "—"
        except Exception:
            return "—"

    def qty_reserved(self, instance: Ticket):
        qty = getattr(instance, "section_qty_reserved", None)
        if qty is None:
            qty = 0
        try:
            from decimal import Decimal
            if isinstance(qty, Decimal):
                return int(qty) if qty == qty.to_integral() else qty
            return int(qty) if float(qty).is_integer() else qty
        except Exception:
            return qty

    def total_euros(self, instance: Ticket):
        euros = getattr(instance, "section_euros_total", None)
        if euros is None:
            euros = 0
        try:
            from decimal import Decimal
            return (Decimal(euros)).quantize(Decimal("1.00"))
        except Exception:
            return 0


class ChildActionsSummaryTable(TableSection):
    verbose_name = _("Action bénévoles")
    height = 240
    related_name = "children_pricesold_for_sections"
    fields = ["price_name", "qty_reserved"]

    def price_name(self, instance: Ticket):
        name = getattr(instance, "section_price_name", None)
        if name:
            return name
        try:
            return instance.reservation.event.name if instance.reservation and instance.reservation.event else "Oups"
        except Exception:
            return "—"

    def qty_reserved(self, instance: Ticket):
        qty = getattr(instance, "section_qty_reserved", None)
        if qty is None:
            qty = 0
        try:
            from decimal import Decimal
            if isinstance(qty, Decimal):
                return int(qty) if qty == qty.to_integral() else qty
            return int(qty) if float(qty).is_integer() else qty
        except Exception:
            return qty

    # Hide the section entirely if the event has no children
    def render(self):
        try:
            if not self.instance.children.exists():
                return ""
        except Exception:
            return ""
        return super().render()


class EventArchiveFilter(admin.SimpleListFilter):
    title = _("Archived")
    parameter_name = "archived"

    def lookups(self, request, model_admin):
        return [
            ("archived", _("Archived")),
        ]

    def queryset(self, request, queryset):
        value = self.value()
        # Filtrage par défaut
        if value is None:
            return queryset.exclude(archived=True)
        if value == "archived":
            return queryset.filter(archived=True)
        return queryset


# Import/Export Resource pour Event
# Resource for CSV import/export of events in admin
class EventResource(resources.ModelResource):
    """Ressource d'import/export pour les événements.
    Resource for import/export of events.

    Clés uniques: (name, datetime) — identifie si on crée ou met à jour.
    Unique keys: (name, datetime) — determines create vs update.

    Les ForeignKey (postal_address) sont exportées en clair (nom lisible).
    ForeignKey fields (postal_address) are exported as human-readable names.

    Les ManyToMany (products, tag) ne sont pas gérés ici.
    ManyToMany fields (products, tag) are not handled here.
    Il faudrait un widget M2MWidget personnalisé pour les gérer.
    A custom M2MWidget would be needed to handle them.
    """

    # postal_address : on exporte/importe le nom de l'adresse au lieu de l'ID
    # postal_address: export/import the address name instead of the raw PK
    postal_address = fields.Field(
        column_name='postal_address',
        attribute='postal_address',
        widget=ForeignKeyWidget(PostalAddress, field='name'),
    )

    class Meta:
        model = Event
        import_id_fields = ('name', 'datetime')
        fields = (
            'name', 'datetime', 'end_datetime', 'jauge_max', 'max_per_user',
            'short_description', 'long_description', 'published', 'archived',
            'private', 'show_time', 'show_gauge', 'slug', 'is_external',
            'full_url', 'postal_address', 'reservation_button_name',
            'minimum_cashless_required',
        )
        # Ordre des colonnes dans le CSV exporté
        # Column order in the exported CSV
        export_order = fields
        widgets = {
            'datetime': {'format': '%Y-%m-%d %H:%M:%S'},
            'end_datetime': {'format': '%Y-%m-%d %H:%M:%S'},
        }
        # Ne pas lever d'erreur sur les lignes invalides, les ignorer
        # Skip invalid rows instead of raising errors
        skip_unchanged = True
        # Afficher un diff des changements avant import
        # Show a diff of changes before import
        report_skipped = True


@admin.register(Event, site=staff_admin_site)
class EventAdmin(ModelAdmin, ImportExportModelAdmin):
    form = EventForm
    compressed_fields = True  # Default: False
    warn_unsaved_form = True  # Default: False
    date_hierarchy = "datetime"
    ordering = ("-datetime",)

    # Import/Export configuration
    resource_classes = [EventResource]
    export_form_class = ExportForm
    import_form_class = ImportForm

    # Unfold sections (expandable rows)
    list_sections = [
        EventPricesSummaryTable,
        ChildActionsSummaryTable,
    ]
    list_per_page = 20

    change_form_template = 'admin/event/change_form.html'
    change_form_before_template = "admin/event/bilan_link_changeform.html"

    inlines = [EventChildrenInline, ]

    actions_row = ["duplicate_day_plus_one", "duplicate_week_plus_one", "duplicate_week_plus_two",
                   "duplicate_month_plus_one", "archive"]

    fieldsets = (
        (None, {
            'fields': (
                'name',
                # 'categorie',
                'datetime',
                'end_datetime',
                'show_time',
                'img',
                'sticker_img',
                'carrousel',
                'short_description',
                'long_description',
                'jauge_max',
                'show_gauge',
                'postal_address',
                'tag',
                'thematique',
            )
        }),
        (_('Bookings'), {
            'fields': (
                # 'easy_reservation',
                'products',
                'max_per_user',
                'reservation_button_name',
                'custom_confirmation_message',
                'refund_deadline',
            ),
        }),
        (_('Publish'), {
            'fields': (
                'published',
                'private',
                'archived',
            ),
        }),
    )

    list_display = [
        'name',
        'display_bilan_link',
        # 'categorie',
        'display_valid_tickets_count',
        'datetime',
        'show_time',
        'published',
    ]

    list_editable = ['published', ]
    readonly_fields = (
        'display_valid_tickets_count',
    )

    search_fields = ['name']
    list_filter = [
        EventArchiveFilter,
        ('datetime', RangeDateTimeFilter),
        'published',
    ]
    list_filter_submit = True

    autocomplete_fields = [
        "tag",
        "thematique",
        "carrousel",

        # Le autocomplete fields + many2many ne permet pas de filtrage facile
        # Pour filter les produits de type billet, regarder le get_search_results dans ProductAdmin
        "products",
    ]

    formfield_overrides = {
        models.TextField: {
            "widget": WysiwygWidget,
        }
    }

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        # Les events action et les events children doivent s'afficher dans un inline
        return (
            queryset
            .exclude(categorie=Event.ACTION)
            .exclude(parent__isnull=False)
            .select_related('postal_address')
            .prefetch_related(
                'tag', 'carrousel', 'products',
                Prefetch(
                    'reservation',
                    queryset=Reservation.objects.select_related('user_commande')
                    .only('pk', 'datetime', 'status', 'user_commande__email', 'event')
                ),
            )
            .annotate(
                valid_tickets_count_annotated=Count(
                    'reservation__tickets',
                    filter=Q(reservation__tickets__status__in=[Ticket.SCANNED, Ticket.NOT_SCANNED]),
                    distinct=True,
                )
            )
        )

    def save_model(self, request, obj: Event, form, change):
        # Sanitize all TextField inputs to avoid XSS via WysiwYG/TextField
        sanitize_textfields(obj)

        super().save_model(request, obj, form, change)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance
        # Fabrication des pricesold event/prix pour pouvoir être selectionné sur le + billet
        # Doit être dans save_related (pas save_model) car les M2M products
        # ne sont disponibles qu'après que Django les a sauvées.
        for product in obj.products.all():
            for price in product.prices.all():
                get_or_create_price_sold(price=price, event=obj)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_add_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return False

    def has_custom_actions_row_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    # ------------------------------------------------------------------
    # URLs et vues personnalisees pour le bilan / Custom URLs and views for the report
    # ------------------------------------------------------------------

    def get_urls(self):
        """
        Ajoute les routes pour le dashboard et le bilan de billetterie.
        / Adds routes for the dashboard and ticketing report.
        """
        from django.urls import path, re_path
        urls = super().get_urls()
        custom_urls = [
            # Dashboard en premier : evite que "dashboard" soit interprete comme un UUID
            # / Dashboard first: prevents "dashboard" from being parsed as a UUID
            path(
                'dashboard/',
                self.admin_site.admin_view(self.vue_dashboard_billetterie),
                name='BaseBillet_event_dashboard',
            ),
            re_path(
                r'^(?P<object_id>[^/]+)/bilan/$',
                self.admin_site.admin_view(self.vue_bilan),
                name='BaseBillet_event_bilan',
            ),
            re_path(
                r'^(?P<object_id>[^/]+)/bilan/pdf/$',
                self.admin_site.admin_view(self.vue_bilan_pdf),
                name='BaseBillet_event_bilan_pdf',
            ),
            re_path(
                r'^(?P<object_id>[^/]+)/bilan/csv/$',
                self.admin_site.admin_view(self.vue_bilan_csv),
                name='BaseBillet_event_bilan_csv',
            ),
            re_path(
                r'^(?P<object_id>[^/]+)/bilan/excel/$',
                self.admin_site.admin_view(self.vue_bilan_excel),
                name='BaseBillet_event_bilan_excel',
            ),
            path(
                'dashboard/export/',
                self.admin_site.admin_view(self.vue_dashboard_export),
                name='BaseBillet_event_dashboard_export',
            ),
        ]
        return custom_urls + urls

    def vue_bilan(self, request, object_id):
        """
        Affiche le bilan de billetterie pour un evenement.
        / Displays the ticketing report for an event.
        """
        import json
        from django.template.response import TemplateResponse
        from django.shortcuts import get_object_or_404
        from BaseBillet.reports import RapportBilletterieService

        event = get_object_or_404(Event, pk=object_id)
        service = RapportBilletterieService(event)

        synthese = service.calculer_synthese()
        courbe_ventes = service.calculer_courbe_ventes()
        scans = service.calculer_scans()

        # Convertir les tranches horaires au format Chart.js {labels, datasets}
        # Le service retourne {labels, data} — on adapte ici.
        # / Convert time slots to Chart.js format {labels, datasets}
        if scans["tranches_horaires"]:
            scans_tranches_chartjs = {
                "labels": scans["tranches_horaires"]["labels"],
                "datasets": [{
                    "label": str(_("Entries")),
                    "data": scans["tranches_horaires"]["data"],
                }],
            }
            scans_tranches_json = json.dumps(scans_tranches_chartjs)
        else:
            scans_tranches_json = "null"

        # Booleens pour masquer les sections vides dans le template
        # / Booleans to hide empty sections in the template
        courbe_a_des_donnees = len(courbe_ventes.get("labels", [])) > 0
        scans_a_des_donnees = scans["scannes"] > 0 or scans["non_scannes"] > 0 or scans["annules"] > 0

        # Calcul des totaux par tarif / Calculate rate totals
        tarifs = service.calculer_ventes_par_tarif()
        totaux_tarifs = None
        if tarifs:
            totaux_tarifs = {
                "vendus": sum(t["vendus"] for t in tarifs),
                "offerts": sum(t["offerts"] for t in tarifs),
                "ca_ttc": sum(t["ca_ttc"] for t in tarifs),
                "ca_ht": sum(t["ca_ht"] for t in tarifs),
                "tva": sum(t["tva"] for t in tarifs),
                "rembourses": sum(t["rembourses"] for t in tarifs),
            }

        contexte = {
            **self.admin_site.each_context(request),
            "event": event,
            "synthese": synthese,
            "courbe_ventes_json": json.dumps(courbe_ventes) if courbe_a_des_donnees else None,
            "has_courbe_ventes": courbe_a_des_donnees,
            "ventes_par_tarif": tarifs,
            "totaux_tarifs": totaux_tarifs,
            "par_moyen_paiement": service.calculer_par_moyen_paiement(),
            "par_canal": service.calculer_par_canal(),
            "scans": scans,
            "has_scans": scans_a_des_donnees,
            "scans_tranches_json": scans_tranches_json,
            "codes_promo": service.calculer_codes_promo(),
            "remboursements": service.calculer_remboursements(),
            "title": f"Bilan — {event.name}",
            "opts": self.model._meta,
            "has_change_permission": self.has_change_permission(request),
        }

        return TemplateResponse(request, "admin/event/bilan.html", contexte)

    def vue_bilan_pdf(self, request, object_id):
        """
        Exporte le bilan de billetterie en PDF A4 paysage (WeasyPrint).
        / Exports the ticketing report as A4 landscape PDF (WeasyPrint).
        LOCALISATION : Administration/admin/events.py
        """
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        from weasyprint import HTML
        from BaseBillet.models import Configuration
        from django.shortcuts import get_object_or_404
        from BaseBillet.reports import RapportBilletterieService

        event = get_object_or_404(Event, pk=object_id)
        service = RapportBilletterieService(event)
        config = Configuration.get_solo()

        # Calcul des totaux par tarif pour le PDF / Calculate rate totals for PDF
        tarifs_pdf = service.calculer_ventes_par_tarif()
        totaux_tarifs_pdf = None
        if tarifs_pdf:
            totaux_tarifs_pdf = {
                "vendus": sum(t["vendus"] for t in tarifs_pdf),
                "offerts": sum(t["offerts"] for t in tarifs_pdf),
                "ca_ttc": sum(t["ca_ttc"] for t in tarifs_pdf),
                "ca_ht": sum(t["ca_ht"] for t in tarifs_pdf),
                "tva": sum(t["tva"] for t in tarifs_pdf),
                "rembourses": sum(t["rembourses"] for t in tarifs_pdf),
            }

        contexte = {
            "event": event,
            "config": config,
            "synthese": service.calculer_synthese(),
            "ventes_par_tarif": tarifs_pdf,
            "totaux_tarifs": totaux_tarifs_pdf,
            "par_moyen_paiement": service.calculer_par_moyen_paiement(),
            "par_canal": service.calculer_par_canal(),
            "scans": service.calculer_scans(),
            "codes_promo": service.calculer_codes_promo(),
            "remboursements": service.calculer_remboursements(),
            "date_generation": timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M"),
        }

        html_string = render_to_string("admin/event/bilan_pdf.html", contexte)
        pdf_bytes = HTML(string=html_string).write_pdf()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        nom_fichier = f"bilan-{event.slug or event.pk}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        return response

    def vue_bilan_csv(self, request, object_id):
        """
        Exporte le bilan de billetterie en CSV structure (delimiteur ;, UTF-8 BOM).
        / Exports the ticketing report as structured CSV (delimiter ;, UTF-8 BOM).
        LOCALISATION : Administration/admin/events.py
        """
        import csv
        from django.http import HttpResponse
        from django.shortcuts import get_object_or_404
        from BaseBillet.reports import RapportBilletterieService

        event = get_object_or_404(Event, pk=object_id)
        service = RapportBilletterieService(event)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        nom_fichier = f"bilan-{event.slug or event.pk}.csv"
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        response.write('\ufeff')  # BOM UTF-8 pour Excel francais / BOM for French Excel

        writer = csv.writer(response, delimiter=';')

        def euros(centimes):
            """Convertit centimes en string euros. / Converts cents to euro string."""
            if centimes is None:
                return "0.00"
            return f"{centimes / 100:.2f}"

        # -- EN-TETE / HEADER --
        writer.writerow(["BILAN DE BILLETTERIE"])
        writer.writerow(["Evenement", event.name])
        writer.writerow(["Date", str(event.datetime)])
        writer.writerow(["Jauge max", event.jauge_max])
        writer.writerow([])

        # -- SYNTHESE / SUMMARY --
        synthese = service.calculer_synthese()
        writer.writerow(["SYNTHESE"])
        writer.writerow(["Billets vendus", synthese["billets_vendus"]])
        writer.writerow(["Billets scannes", synthese["billets_scannes"]])
        writer.writerow(["No-show", synthese["no_show"]])
        writer.writerow(["CA TTC", euros(synthese["ca_ttc"])])
        writer.writerow(["Remboursements", euros(synthese["remboursements"])])
        writer.writerow(["CA net", euros(synthese["ca_net"])])
        writer.writerow(["Taux remplissage", f"{synthese['taux_remplissage']}%"])
        writer.writerow([])

        # -- VENTES PAR TARIF / SALES BY RATE --
        tarifs = service.calculer_ventes_par_tarif()
        writer.writerow(["VENTES PAR TARIF"])
        writer.writerow(["Tarif", "Vendus", "Offerts", "CA TTC", "HT", "TVA", "Rembourses"])
        for t in tarifs:
            writer.writerow([
                t["nom"], t["vendus"], t["offerts"],
                euros(t["ca_ttc"]), euros(t["ca_ht"]), euros(t["tva"]),
                t["rembourses"],
            ])
        # Ligne TOTAL / TOTAL row
        if tarifs:
            writer.writerow([
                "TOTAL",
                sum(t["vendus"] for t in tarifs),
                sum(t["offerts"] for t in tarifs),
                euros(sum(t["ca_ttc"] for t in tarifs)),
                euros(sum(t["ca_ht"] for t in tarifs)),
                euros(sum(t["tva"] for t in tarifs)),
                sum(t["rembourses"] for t in tarifs),
            ])
        writer.writerow([])

        # -- PAR MOYEN DE PAIEMENT / BY PAYMENT METHOD --
        moyens = service.calculer_par_moyen_paiement()
        writer.writerow(["PAR MOYEN DE PAIEMENT"])
        writer.writerow(["Moyen", "Montant", "Pourcentage", "Nb billets"])
        for m in moyens:
            writer.writerow([
                m["label"], euros(m["montant"]),
                f"{m['pourcentage']}%", m["nb_billets"],
            ])
        writer.writerow([])

        # -- PAR CANAL DE VENTE / BY SALES CHANNEL (conditionnel) --
        canaux = service.calculer_par_canal()
        if canaux:
            writer.writerow(["PAR CANAL DE VENTE"])
            writer.writerow(["Canal", "Nb billets", "Montant"])
            for c in canaux:
                writer.writerow([c["label"], c["nb_billets"], euros(c["montant"])])
            writer.writerow([])

        # -- SCANS --
        scans = service.calculer_scans()
        writer.writerow(["SCANS"])
        writer.writerow(["Scannes", scans["scannes"]])
        writer.writerow(["Non scannes", scans["non_scannes"]])
        writer.writerow(["Annules", scans["annules"]])
        writer.writerow([])

        # -- CODES PROMO / PROMO CODES (conditionnel) --
        promos = service.calculer_codes_promo()
        if promos:
            writer.writerow(["CODES PROMO"])
            writer.writerow(["Code", "Utilisations", "Reduction", "Manque a gagner"])
            for p in promos:
                writer.writerow([
                    p["nom"], p["utilisations"],
                    f"{p['taux_reduction']}%", euros(p["manque_a_gagner"]),
                ])
            writer.writerow([])

        # -- REMBOURSEMENTS / REFUNDS --
        remb = service.calculer_remboursements()
        writer.writerow(["REMBOURSEMENTS"])
        writer.writerow(["Nombre", remb["nombre"]])
        writer.writerow(["Montant total", euros(remb["montant_total"])])
        writer.writerow(["Taux", f"{remb['taux']}%"])

        return response

    def vue_bilan_excel(self, request, object_id):
        """
        Exporte le bilan de billetterie en Excel (.xlsx, openpyxl).
        / Exports the ticketing report as Excel (.xlsx, openpyxl).
        LOCALISATION : Administration/admin/events.py
        """
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from django.shortcuts import get_object_or_404
        from BaseBillet.reports import RapportBilletterieService

        event = get_object_or_404(Event, pk=object_id)
        service = RapportBilletterieService(event)

        def euros(centimes):
            """Convertit centimes en float euros. / Converts cents to float euros."""
            if centimes is None:
                return 0.0
            return centimes / 100

        # Styles
        section_font = Font(bold=True, size=11, color="FFFFFF")
        section_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        euro_format = '#,##0.00 "EUR"'
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bilan"

        row = 1

        # -- Fonction utilitaire pour ecrire une ligne de section --
        def write_section(titre):
            nonlocal row
            cell = ws.cell(row=row, column=1, value=titre)
            cell.font = section_font
            cell.fill = section_fill
            for col in range(2, 8):
                ws.cell(row=row, column=col).fill = section_fill
            row += 1

        def write_header(colonnes):
            nonlocal row
            for i, col_name in enumerate(colonnes, 1):
                cell = ws.cell(row=row, column=i, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

        def write_row(valeurs, formats=None):
            nonlocal row
            for i, val in enumerate(valeurs, 1):
                cell = ws.cell(row=row, column=i, value=val)
                cell.border = thin_border
                if formats and i <= len(formats) and formats[i - 1]:
                    cell.number_format = formats[i - 1]
            row += 1

        # EN-TETE / HEADER
        write_section("BILAN DE BILLETTERIE")
        write_row(["Evenement", event.name])
        write_row(["Date", str(event.datetime)])
        write_row(["Jauge max", event.jauge_max])
        row += 1

        # SYNTHESE / SUMMARY
        synthese = service.calculer_synthese()
        write_section("SYNTHESE")
        write_row(["Billets vendus", synthese["billets_vendus"]])
        write_row(["Billets scannes", synthese["billets_scannes"]])
        write_row(["No-show", synthese["no_show"]])
        write_row(["CA TTC", euros(synthese["ca_ttc"])], [None, euro_format])
        write_row(["Remboursements", euros(synthese["remboursements"])], [None, euro_format])
        write_row(["CA net", euros(synthese["ca_net"])], [None, euro_format])
        write_row(["Taux remplissage", f"{synthese['taux_remplissage']}%"])
        row += 1

        # VENTES PAR TARIF / SALES BY RATE
        tarifs = service.calculer_ventes_par_tarif()
        write_section("VENTES PAR TARIF")
        write_header(["Tarif", "Vendus", "Offerts", "CA TTC", "HT", "TVA", "Rembourses"])
        fmt_tarif = [None, None, None, euro_format, euro_format, euro_format, None]
        for t in tarifs:
            write_row([
                t["nom"], t["vendus"], t["offerts"],
                euros(t["ca_ttc"]), euros(t["ca_ht"]), euros(t["tva"]),
                t["rembourses"],
            ], fmt_tarif)
        if tarifs:
            write_row([
                "TOTAL",
                sum(t["vendus"] for t in tarifs),
                sum(t["offerts"] for t in tarifs),
                euros(sum(t["ca_ttc"] for t in tarifs)),
                euros(sum(t["ca_ht"] for t in tarifs)),
                euros(sum(t["tva"] for t in tarifs)),
                sum(t["rembourses"] for t in tarifs),
            ], fmt_tarif)
        row += 1

        # PAR MOYEN DE PAIEMENT / BY PAYMENT METHOD
        moyens = service.calculer_par_moyen_paiement()
        write_section("PAR MOYEN DE PAIEMENT")
        write_header(["Moyen", "Montant", "Pourcentage", "Nb billets"])
        for m in moyens:
            write_row([
                m["label"], euros(m["montant"]),
                f"{m['pourcentage']}%", m["nb_billets"],
            ], [None, euro_format, None, None])
        row += 1

        # PAR CANAL DE VENTE / BY SALES CHANNEL
        canaux = service.calculer_par_canal()
        if canaux:
            write_section("PAR CANAL DE VENTE")
            write_header(["Canal", "Nb billets", "Montant"])
            for c in canaux:
                write_row([c["label"], c["nb_billets"], euros(c["montant"])],
                          [None, None, euro_format])
            row += 1

        # SCANS
        scans = service.calculer_scans()
        write_section("SCANS")
        write_row(["Scannes", scans["scannes"]])
        write_row(["Non scannes", scans["non_scannes"]])
        write_row(["Annules", scans["annules"]])
        row += 1

        # CODES PROMO / PROMO CODES
        promos = service.calculer_codes_promo()
        if promos:
            write_section("CODES PROMO")
            write_header(["Code", "Utilisations", "Reduction", "Manque a gagner"])
            for p in promos:
                write_row([
                    p["nom"], p["utilisations"],
                    f"{p['taux_reduction']}%", euros(p["manque_a_gagner"]),
                ], [None, None, None, euro_format])
            row += 1

        # REMBOURSEMENTS / REFUNDS
        remb = service.calculer_remboursements()
        write_section("REMBOURSEMENTS")
        write_row(["Nombre", remb["nombre"]])
        write_row(["Montant total", euros(remb["montant_total"])], [None, euro_format])
        write_row(["Taux", f"{remb['taux']}%"])

        # Ajuster la largeur des colonnes / Adjust column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14

        # Generer la reponse HTTP / Generate HTTP response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        nom_fichier = f"bilan-{event.slug or event.pk}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        return response

    # ------------------------------------------------------------------
    # Export par periode / Period export
    # ------------------------------------------------------------------

    def _query_events_periode(self, date_debut, date_fin):
        """
        Query annotee des events pour une periode donnee (filtrage sur datetime__date).
        Retourne un queryset d'Event avec annotations nb_vendus, nb_scannes, ca_ttc, ca_rembourse.
        / Annotated query for events in a period (filter on datetime__date).
        Returns an Event queryset with annotations nb_vendus, nb_scannes, ca_ttc, ca_rembourse.

        LOCALISATION : Administration/admin/events.py
        """
        from django.db.models import Sum
        from django.db.models.functions import Coalesce

        return Event.objects.filter(
            archived=False,
            parent__isnull=True,
            datetime__date__gte=date_debut,
            datetime__date__lte=date_fin,
        ).exclude(
            categorie=Event.ACTION,
        ).annotate(
            nb_vendus=Count(
                'reservation__tickets',
                filter=Q(reservation__tickets__status__in=[Ticket.NOT_SCANNED, Ticket.SCANNED]),
                distinct=True,
            ),
            nb_scannes=Count(
                'reservation__tickets',
                filter=Q(reservation__tickets__status=Ticket.SCANNED),
                distinct=True,
            ),
            ca_ttc=Coalesce(
                Sum('reservation__lignearticles__amount',
                    filter=Q(reservation__lignearticles__status=LigneArticle.VALID)),
                0,
            ),
            ca_rembourse=Coalesce(
                Sum('reservation__lignearticles__amount',
                    filter=Q(reservation__lignearticles__status=LigneArticle.REFUNDED)),
                0,
            ),
        ).select_related('postal_address').order_by('datetime')

    def _calculer_synthese_periode(self, events):
        """
        Calcule les totaux globaux a partir d'un queryset d'events annotes.
        / Computes global totals from an annotated event queryset.

        LOCALISATION : Administration/admin/events.py
        """
        billets_vendus = sum(e.nb_vendus for e in events)
        billets_scannes = sum(e.nb_scannes for e in events)
        ca_ttc = sum(e.ca_ttc for e in events)
        ca_rembourse = sum(e.ca_rembourse for e in events)
        ca_net = ca_ttc - ca_rembourse

        return {
            "billets_vendus": billets_vendus,
            "billets_scannes": billets_scannes,
            "ca_ttc": ca_ttc,
            "ca_rembourse": ca_rembourse,
            "ca_net": ca_net,
        }

    def vue_dashboard_export(self, request):
        """
        GET : retourne le formulaire partial (pour hx-get).
        POST : valide les dates, genere CSV/PDF/Excel multi-events.
        / GET: returns the partial form (for hx-get).
        POST: validates dates, generates CSV/PDF/Excel multi-event export.

        LOCALISATION : Administration/admin/events.py
        """
        from django.http import HttpResponse
        from django.template.response import TemplateResponse

        # GET : retourner le formulaire HTMX / GET: return the HTMX form
        if request.method == 'GET':
            contexte = {
                "form_action_url": reverse('staff_admin:BaseBillet_event_dashboard_export'),
            }
            return TemplateResponse(
                request,
                "admin/event/partials/export_periode_form.html",
                contexte,
            )

        # POST : valider et exporter / POST: validate and export
        serializer = ExportPeriodeSerializer(data=request.POST)
        if not serializer.is_valid():
            from django.contrib import messages as django_messages
            for field, erreurs in serializer.errors.items():
                for erreur in erreurs:
                    django_messages.error(request, f"{field}: {erreur}")
            return redirect(reverse('staff_admin:BaseBillet_event_dashboard'))

        date_debut = serializer.validated_data['date_debut']
        date_fin = serializer.validated_data['date_fin']
        format_export = serializer.validated_data['format_export']

        # Query annotee fraiche (pas de cache) / Fresh annotated query (no cache)
        events_qs = self._query_events_periode(date_debut, date_fin)
        events_liste = list(events_qs)

        # Enrichir chaque event avec ca_net / Enrich each event with ca_net
        for event in events_liste:
            event.ca_net = event.ca_ttc - event.ca_rembourse

        synthese = self._calculer_synthese_periode(events_liste)

        if format_export == 'csv':
            return self._export_periode_csv(events_liste, synthese, date_debut, date_fin)
        elif format_export == 'pdf':
            return self._export_periode_pdf(events_liste, synthese, date_debut, date_fin)
        elif format_export == 'excel':
            return self._export_periode_excel(events_liste, synthese, date_debut, date_fin)

    def _export_periode_csv(self, events, synthese, date_debut, date_fin):
        """
        Genere le CSV multi-events pour une periode.
        / Generates the multi-event CSV for a period.
        """
        import csv
        from django.http import HttpResponse

        def euros(centimes):
            if centimes is None:
                return "0.00"
            return f"{centimes / 100:.2f}"

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        nom_fichier = f"bilan-periode-{date_debut}-{date_fin}.csv"
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        response.write('\ufeff')  # BOM UTF-8

        writer = csv.writer(response, delimiter=';')

        # EN-TETE / HEADER
        writer.writerow(["BILAN BILLETTERIE — PERIODE"])
        writer.writerow(["Du", str(date_debut)])
        writer.writerow(["Au", str(date_fin)])
        writer.writerow(["Nb evenements", len(events)])
        writer.writerow([])

        # SYNTHESE GLOBALE / GLOBAL SUMMARY
        writer.writerow(["SYNTHESE GLOBALE"])
        writer.writerow(["Billets vendus", synthese["billets_vendus"]])
        writer.writerow(["Billets scannes", synthese["billets_scannes"]])
        writer.writerow(["CA TTC", euros(synthese["ca_ttc"])])
        writer.writerow(["Remboursements", euros(synthese["ca_rembourse"])])
        writer.writerow(["CA net", euros(synthese["ca_net"])])
        writer.writerow([])

        # DETAIL PAR EVENEMENT / DETAIL BY EVENT
        writer.writerow(["DETAIL PAR EVENEMENT"])
        writer.writerow(["Evenement", "Date", "Vendus", "Scannes", "CA TTC", "Remboursements", "CA net"])
        for e in events:
            writer.writerow([
                e.name,
                str(e.datetime),
                e.nb_vendus,
                e.nb_scannes,
                euros(e.ca_ttc),
                euros(e.ca_rembourse),
                euros(e.ca_net),
            ])

        # TOTAL
        if events:
            writer.writerow([
                "TOTAL", "",
                synthese["billets_vendus"],
                synthese["billets_scannes"],
                euros(synthese["ca_ttc"]),
                euros(synthese["ca_rembourse"]),
                euros(synthese["ca_net"]),
            ])

        return response

    def _export_periode_pdf(self, events, synthese, date_debut, date_fin):
        """
        Genere le PDF multi-events pour une periode (WeasyPrint).
        / Generates the multi-event PDF for a period (WeasyPrint).
        """
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        from weasyprint import HTML

        config = Configuration.get_solo()

        contexte = {
            "events": events,
            "synthese": synthese,
            "date_debut": date_debut,
            "date_fin": date_fin,
            "nb_events": len(events),
            "config": config,
            "date_generation": timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M"),
        }

        html_string = render_to_string("admin/event/bilan_periode_pdf.html", contexte)
        pdf_bytes = HTML(string=html_string).write_pdf()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        nom_fichier = f"bilan-periode-{date_debut}-{date_fin}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        return response

    def _export_periode_excel(self, events, synthese, date_debut, date_fin):
        """
        Genere le Excel multi-events pour une periode (openpyxl, 2 onglets).
        / Generates the multi-event Excel for a period (openpyxl, 2 sheets).
        """
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        from django.http import HttpResponse

        def euros(centimes):
            if centimes is None:
                return 0.0
            return centimes / 100

        euro_format = '#,##0.00 "EUR"'

        # Styles
        section_font = Font(bold=True, size=11, color="FFFFFF")
        section_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        wb = openpyxl.Workbook()

        # --- Onglet 1 : Synthese / Sheet 1: Summary ---
        ws_synthese = wb.active
        ws_synthese.title = "Synthese"

        row = 1

        def write_section_s(titre):
            nonlocal row
            cell = ws_synthese.cell(row=row, column=1, value=titre)
            cell.font = section_font
            cell.fill = section_fill
            for col in range(2, 8):
                ws_synthese.cell(row=row, column=col).fill = section_fill
            row += 1

        def write_row_s(valeurs, formats=None):
            nonlocal row
            for i, val in enumerate(valeurs, 1):
                cell = ws_synthese.cell(row=row, column=i, value=val)
                cell.border = thin_border
                if formats and i <= len(formats) and formats[i - 1]:
                    cell.number_format = formats[i - 1]
            row += 1

        write_section_s("BILAN BILLETTERIE — PERIODE")
        write_row_s(["Du", str(date_debut)])
        write_row_s(["Au", str(date_fin)])
        write_row_s(["Nb evenements", len(events)])
        row += 1

        write_section_s("SYNTHESE GLOBALE")
        write_row_s(["Billets vendus", synthese["billets_vendus"]])
        write_row_s(["Billets scannes", synthese["billets_scannes"]])
        write_row_s(["CA TTC", euros(synthese["ca_ttc"])], [None, euro_format])
        write_row_s(["Remboursements", euros(synthese["ca_rembourse"])], [None, euro_format])
        write_row_s(["CA net", euros(synthese["ca_net"])], [None, euro_format])

        ws_synthese.column_dimensions['A'].width = 22
        ws_synthese.column_dimensions['B'].width = 18

        # --- Onglet 2 : Detail par evenement / Sheet 2: Detail by event ---
        ws_detail = wb.create_sheet("Detail par evenement")
        row_d = 1

        # En-tete / Header
        colonnes = ["Evenement", "Date", "Vendus", "Scannes", "CA TTC", "Remboursements", "CA net"]
        for i, col_name in enumerate(colonnes, 1):
            cell = ws_detail.cell(row=row_d, column=i, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row_d += 1

        fmt_detail = [None, None, None, None, euro_format, euro_format, euro_format]
        for e in events:
            valeurs = [
                e.name,
                e.datetime.strftime("%d/%m/%Y %H:%M") if e.datetime else "",
                e.nb_vendus,
                e.nb_scannes,
                euros(e.ca_ttc),
                euros(e.ca_rembourse),
                euros(e.ca_net),
            ]
            for i, val in enumerate(valeurs, 1):
                cell = ws_detail.cell(row=row_d, column=i, value=val)
                cell.border = thin_border
                if fmt_detail[i - 1]:
                    cell.number_format = fmt_detail[i - 1]
            row_d += 1

        # Ligne TOTAL / TOTAL row
        if events:
            total_row = [
                "TOTAL", "",
                synthese["billets_vendus"],
                synthese["billets_scannes"],
                euros(synthese["ca_ttc"]),
                euros(synthese["ca_rembourse"]),
                euros(synthese["ca_net"]),
            ]
            for i, val in enumerate(total_row, 1):
                cell = ws_detail.cell(row=row_d, column=i, value=val)
                cell.font = Font(bold=True)
                cell.border = thin_border
                if fmt_detail[i - 1]:
                    cell.number_format = fmt_detail[i - 1]

        ws_detail.column_dimensions['A'].width = 30
        ws_detail.column_dimensions['B'].width = 18
        ws_detail.column_dimensions['C'].width = 12
        ws_detail.column_dimensions['D'].width = 12
        ws_detail.column_dimensions['E'].width = 14
        ws_detail.column_dimensions['F'].width = 16
        ws_detail.column_dimensions['G'].width = 14

        # Generer la reponse / Generate response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        nom_fichier = f"bilan-periode-{date_debut}-{date_fin}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        return response

    # ------------------------------------------------------------------
    # Dashboard billetterie / Ticketing dashboard
    # ------------------------------------------------------------------

    def vue_dashboard_billetterie(self, request):
        """
        Dashboard billetterie : vue d'ensemble des events avec indicateurs.
        Affiche les events a venir (tous) et les 6 derniers passes.
        Utilise une query annotee unique + cache 2 minutes.
        / Ticketing dashboard: event overview with key indicators.

        LOCALISATION : Administration/admin/events.py
        """
        from django.core.cache import cache
        from django.db.models import Sum
        from django.db.models.functions import Coalesce
        from django.template.response import TemplateResponse
        from BaseBillet.models import LigneArticle, Ticket

        cache_key = f"dashboard_billetterie:{connection.tenant.pk}"
        donnees_cachees = cache.get(cache_key)

        if donnees_cachees is None:
            maintenant = timezone.now()

            # Query annotee unique : 1 seule requete pour tous les events + indicateurs
            # / Single annotated query: 1 query for all events + indicators
            events_avec_stats = Event.objects.filter(
                archived=False,
                parent__isnull=True,
            ).exclude(
                categorie=Event.ACTION,
            ).annotate(
                nb_vendus=Count(
                    'reservation__tickets',
                    filter=Q(reservation__tickets__status__in=[Ticket.NOT_SCANNED, Ticket.SCANNED]),
                    distinct=True,
                ),
                nb_scannes=Count(
                    'reservation__tickets',
                    filter=Q(reservation__tickets__status=Ticket.SCANNED),
                    distinct=True,
                ),
                ca_ttc=Coalesce(
                    Sum('reservation__lignearticles__amount',
                        filter=Q(reservation__lignearticles__status=LigneArticle.VALID)),
                    0,
                ),
                ca_rembourse=Coalesce(
                    Sum('reservation__lignearticles__amount',
                        filter=Q(reservation__lignearticles__status=LigneArticle.REFUNDED)),
                    0,
                ),
                nb_reservations=Count('reservation', distinct=True),
            ).select_related('postal_address')

            # Separer events a venir et passes
            # / Separate upcoming and past events
            events_a_venir_qs = events_avec_stats.filter(
                datetime__gte=maintenant,
            ).order_by('datetime')

            events_passes_qs = events_avec_stats.filter(
                datetime__lt=maintenant,
            ).order_by('-datetime')[:6]

            # Calculer ca_net et taux_remplissage en Python pour chaque event
            # / Compute ca_net and fill_rate in Python for each event
            def enrichir_event(event):
                event.ca_net = event.ca_ttc - event.ca_rembourse
                if event.jauge_max and event.jauge_max > 0:
                    event.taux_remplissage = round((event.nb_vendus / event.jauge_max) * 100, 1)
                else:
                    event.taux_remplissage = 0.0
                return event

            events_a_venir = [enrichir_event(e) for e in events_a_venir_qs]
            events_passes = [enrichir_event(e) for e in events_passes_qs]

            donnees_cachees = {
                "events_a_venir": events_a_venir,
                "events_passes": events_passes,
            }
            cache.set(cache_key, donnees_cachees, 120)  # TTL 2 minutes

        contexte = {
            **self.admin_site.each_context(request),
            "events_a_venir": donnees_cachees["events_a_venir"],
            "events_passes": donnees_cachees["events_passes"],
            "title": _("Ticketing dashboard"),
            "opts": self.model._meta,
        }

        return TemplateResponse(request, "admin/event/dashboard_billetterie.html", contexte)

    # ------------------------------------------------------------------
    # Colonnes d'affichage / Display columns
    # ------------------------------------------------------------------

    @display(description=_("Report"))
    def display_bilan_link(self, obj):
        """
        Affiche un lien vers le bilan si l'event a des reservations.
        / Displays a link to the report if the event has reservations.
        """
        from django.utils.html import format_html

        # Utiliser l'annotation si disponible, sinon requete directe
        # / Use annotation if available, otherwise direct query
        nombre_reservations = Reservation.objects.filter(event=obj).count()
        if nombre_reservations == 0:
            return "—"

        url = reverse('staff_admin:BaseBillet_event_bilan', args=[obj.pk])
        return format_html(
            '<a href="{}" title="{}" data-testid="bilan-link">'
            '<span class="material-symbols-outlined" style="font-size: 20px;" aria-hidden="true">assessment</span>'
            '</a>',
            url, _("View report"),
        )

    @display(description=_("Valid tickets"))
    def display_valid_tickets_count(self, instance: Event):
        # Use annotated value to avoid N+1; fallback to method if not present (e.g., detail page)
        count = getattr(instance, 'valid_tickets_count_annotated', None)
        if count is None:
            count = instance.valid_tickets_count()
        return f"{count} / {instance.jauge_max}"

    @action(
        description=_("Archive"),
        permissions=["custom_actions_row"],
    )
    def archive(self, request, object_id):
        event = Event.objects.get(pk=object_id)
        event.archived = True
        event.published = False
        event.save(update_fields=['archived', 'published'])
        return redirect(request.META["HTTP_REFERER"])

    @action(
        description=_("Duplicate (day+1)"),
        permissions=["custom_actions_row"],
    )
    def duplicate_day_plus_one(self, request, object_id):
        """Duplicate an event with the date set to the next day"""
        obj = Event.objects.get(pk=object_id)
        try:
            duplicate = self._duplicate_event(obj, date_adjustment="day")
            messages.success(request, _("Event duplicated successfully"))
        except IntegrityError as e:
            messages.error(request, _("Un evenement avec le même nom et date semble déja dupliqué"))

        return redirect(request.META["HTTP_REFERER"])

    @action(
        description=_("Duplicate (week+1)"),
        permissions=["custom_actions_row"],
    )
    def duplicate_week_plus_one(self, request, object_id):
        """Duplicate an event with the date set to the next week"""
        obj = Event.objects.get(pk=object_id)
        try:
            duplicate = self._duplicate_event(obj, date_adjustment="week")
            messages.success(request, _("Event duplicated successfully"))
        except IntegrityError as e:
            messages.error(request, _("Un evenement avec le même nom et date semble déja dupliqué"))

        return redirect(request.META["HTTP_REFERER"])

    @action(
        description=_("Duplicate (week+2)"),
        permissions=["custom_actions_row"],
    )
    def duplicate_week_plus_two(self, request, object_id):
        """Duplicate an event with the date set to two weeks ahead"""
        obj = Event.objects.get(pk=object_id)
        try:
            duplicate = self._duplicate_event(obj, date_adjustment="week2")
            messages.success(request, _("Event duplicated successfully"))
        except IntegrityError as e:
            messages.error(request, _("Un evenement avec le même nom et date semble déja dupliqué"))

        return redirect(request.META["HTTP_REFERER"])

    @action(
        description=_("Duplicate (month+1)"),
        permissions=["custom_actions_row"],
    )
    def duplicate_month_plus_one(self, request, object_id):
        """Duplicate an event with the date set to the next month"""
        obj = Event.objects.get(pk=object_id)
        try:
            duplicate = self._duplicate_event(obj, date_adjustment="month")
            messages.success(request, _("Event duplicated successfully"))
        except IntegrityError as e:
            messages.error(request, _("Un evenement avec le même nom et date semble déja dupliqué"))
        return redirect(request.META["HTTP_REFERER"])

    def _duplicate_event(self, obj, date_adjustment=None):
        """
        Helper method to duplicate an event

        Args:
            obj: The event to duplicate
            date_adjustment: Type of date adjustment to apply ("day", "week", "month", or None for same date)

        Returns:
            The duplicated event
        """
        # Create a copy of the event
        duplicate = Event.objects.get(uuid=obj.uuid)
        duplicate.pk = None  # This will create a new object on save
        duplicate.rsa_key = None  # Ensure a new RSA key is generated
        duplicate.slug = None  # Ensure a new slug is generated

        # Set the name (no prefix)
        duplicate.name = obj.name

        # Set published to False
        duplicate.published = False

        # Adjust the date based on the date_adjustment parameter
        if date_adjustment == "day":
            # Add 1 day to the date
            duplicate.datetime = obj.datetime + timedelta(days=1)
            if obj.end_datetime:
                duplicate.end_datetime = obj.end_datetime + timedelta(days=1)
        elif date_adjustment == "week":
            # Add 7 days to the date
            duplicate.datetime = obj.datetime + timedelta(days=7)
            if obj.end_datetime:
                duplicate.end_datetime = obj.end_datetime + timedelta(days=7)
        elif date_adjustment == "week2":
            # Add 14 days to the date
            duplicate.datetime = obj.datetime + timedelta(days=14)
            if obj.end_datetime:
                duplicate.end_datetime = obj.end_datetime + timedelta(days=14)
        elif date_adjustment == "month":
            # Add 1 month to the date
            from dateutil.relativedelta import relativedelta
            duplicate.datetime = obj.datetime + relativedelta(months=1)
            if obj.end_datetime:
                duplicate.end_datetime = obj.end_datetime + relativedelta(months=1)

        # Save the duplicate
        duplicate.save()

        # Copy many-to-many relationships
        duplicate.products.set(obj.products.all())
        duplicate.tag.set(obj.tag.all())
        duplicate.carrousel.set(obj.carrousel.all())

        # Duplicate child events of type ACTION
        for child in obj.children.filter(categorie=Event.ACTION):
            child_duplicate = Event.objects.get(uuid=child.uuid)
            child_duplicate.pk = None  # This will create a new object on save
            child_duplicate.rsa_key = None  # Ensure a new RSA key is generated
            child_duplicate.slug = None  # Ensure a new slug is generated
            child_duplicate.parent = duplicate

            # Child events should be published
            child_duplicate.published = True

            # Adjust the date based on the date_adjustment parameter
            if date_adjustment == "day":
                # Add 1 day to the date
                child_duplicate.datetime = child.datetime + timedelta(days=1)
                if child.end_datetime:
                    child_duplicate.end_datetime = child.end_datetime + timedelta(days=1)
            elif date_adjustment == "week":
                # Add 7 days to the date
                child_duplicate.datetime = child.datetime + timedelta(days=7)
                if child.end_datetime:
                    child_duplicate.end_datetime = child.end_datetime + timedelta(days=7)
            elif date_adjustment == "week2":
                # Add 14 days to the date
                child_duplicate.datetime = child.datetime + timedelta(days=14)
                if child.end_datetime:
                    child_duplicate.end_datetime = child.end_datetime + timedelta(days=14)
            elif date_adjustment == "month":
                # Add 1 month to the date
                from dateutil.relativedelta import relativedelta
                child_duplicate.datetime = child.datetime + relativedelta(months=1)
                if child.end_datetime:
                    child_duplicate.end_datetime = child.end_datetime + relativedelta(months=1)

            child_duplicate.save()

            # Copy many-to-many relationships for child
            child_duplicate.products.set(child.products.all())
            child_duplicate.tag.set(child.tag.all())
            child_duplicate.carrousel.set(child.carrousel.all())

        return duplicate
