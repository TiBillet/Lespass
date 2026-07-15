"""
Administration des modeles LaBoutik (caisse, points de vente, imprimantes, tables, commandes).
/ Admin for LaBoutik models (POS, points of sale, printers, tables, orders).

LOCALISATION : Administration/admin/laboutik.py
"""
import logging

from django import forms
from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.db import connection
from django.shortcuts import get_object_or_404
from django.utils.html import format_html
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from solo.admin import SingletonModelAdmin
from unfold.admin import ModelAdmin, TabularInline

from Administration.admin.products import ICON_POS, IconPickerWidget
from Administration.admin.site import staff_admin_site
from ApiBillet.permissions import TenantAdminPermissionWithRequest
from QrcodeCashless.models import CarteCashless
from laboutik.models import (
    LaboutikConfiguration,
    Printer,
    Terminal,
    TPEBancaire,
    PointDeVente, CartePrimaire, CategorieTable, Table,
    CommandeSauvegarde, ArticleCommandeSauvegarde,
    ClotureCaisse,
    ImpressionLog,
    JournalOperation,
    HistoriqueFondDeCaisse,
    CompteComptable,
    MappingMoyenDePaiement,
)

logger = logging.getLogger(__name__)


@admin.register(LaboutikConfiguration, site=staff_admin_site)
class LaboutikConfigurationAdmin(SingletonModelAdmin, ModelAdmin):
    """Admin singleton pour la configuration globale de l'interface caisse.
    Singleton admin for the global POS interface configuration.
    LOCALISATION : Administration/admin/laboutik.py"""
    compressed_fields = True
    warn_unsaved_form = True

    # Le compteur de tickets est en lecture seule pour eviter une remise a zero accidentelle.
    # / Receipt counter is read-only to prevent accidental reset.
    readonly_fields = ('compteur_tickets',)

    fieldsets = (
        (_('Interface caisse / POS interface'), {
            'fields': (
                'taille_police_articles',
                'mode_ecole',
            ),
        }),
        (_('Sunmi Cloud'), {
            'fields': (
                'sunmi_app_id',
                'sunmi_app_key',
            ),
            'description': _(
                "Identifiants Sunmi Cloud (stockes chiffres). "
                "/ Sunmi Cloud credentials (stored encrypted)."
            ),
        }),
        (_('Ticket de vente / Sale receipt'), {
            'fields': (
                'pied_ticket',
                'compteur_tickets',
            ),
            'description': _(
                "Personnalisation des tickets de vente. "
                "/ Sale receipt customization."
            ),
        }),
        (_('Rapports automatiques / Automatic reports'), {
            'fields': (
                'rapport_emails',
                'rapport_periodicite',
            ),
            'description': _(
                "Envoi automatique des rapports de cloture par email (7h locale). "
                "/ Automatic closure report email sending (7am local time)."
            ),
        }),
    )

    def has_add_permission(self, request):
        # Singleton : pas de creation manuelle — get_or_create suffit
        # Singleton: no manual creation — get_or_create is enough
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


class PointDeVenteForm(forms.ModelForm):
    """Formulaire pour les points de vente avec selecteur visuel d'icones.
    Form for points of sale with visual icon picker.
    LOCALISATION : Administration/admin/laboutik.py"""

    icon = forms.ChoiceField(
        choices=[("", _("— Aucune icône —"))] + list(ICON_POS),
        required=False,
        label=_("Icon"),
        widget=IconPickerWidget(),
    )

    class Meta:
        model = PointDeVente
        fields = '__all__'


@admin.register(PointDeVente, site=staff_admin_site)
class PointDeVenteAdmin(ModelAdmin):
    """Admin pour les points de vente.
    Admin for points of sale.
    LOCALISATION : Administration/admin/laboutik.py"""
    form = PointDeVenteForm
    compressed_fields = True
    warn_unsaved_form = True

    list_display = ('name', 'comportement', 'service_direct', 'hidden')
    list_filter = ['comportement', 'hidden']
    search_fields = ['name']
    ordering = ('poid_liste', 'name')
    filter_horizontal = ('products', 'categories')

    fieldsets = (
        (_('General'), {
            'fields': (
                'name',
                'icon',
                'comportement',
                'poid_liste',
                'hidden',
            ),
        }),
        (_('Options'), {
            'fields': (
                'service_direct',
                'afficher_les_prix',
                'accepte_especes',
                'accepte_carte_bancaire',
                'accepte_cheque',
                'accepte_commandes',
            ),
        }),
        (_('Products & categories'), {
            'fields': (
                'products',
                'categories',
            ),
        }),
    )

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


@admin.register(Printer, site=staff_admin_site)
class PrinterAdmin(ModelAdmin):
    """Admin pour les imprimantes thermiques (Sunmi Cloud / Inner / LAN).
    Admin for thermal printers (Sunmi Cloud / Inner / LAN).
    LOCALISATION : Administration/admin/laboutik.py"""
    compressed_fields = True
    warn_unsaved_form = True

    list_display = (
        'name', 'printer_type', 'dots_per_line', 'sunmi_serial_number',
        'terminaux_qui_impriment', 'active',
    )
    list_filter = ['printer_type', 'active']
    search_fields = ['name', 'sunmi_serial_number']
    ordering = ('name',)

    @admin.display(description=_("Terminaux"))
    def terminaux_qui_impriment(self, printer):
        """
        Les terminaux qui sortent leurs tickets sur cette imprimante.
        Plusieurs terminaux peuvent partager la meme imprimante.
        / The terminals that print their tickets on this printer.
        """
        noms_des_terminaux = [
            terminal.name or str(terminal.id)
            for terminal in printer.terminaux.all()
        ]
        if not noms_des_terminaux:
            return "—"
        return ", ".join(noms_des_terminaux)

    def get_queryset(self, request):
        # Prefetch pour eviter une requete par imprimante dans la colonne ci-dessus.
        # / Prefetch to avoid one query per printer in the column above.
        return super().get_queryset(request).prefetch_related('terminaux')

    actions = ['imprimer_un_ticket_de_test']

    @admin.action(description=_("Imprimer un ticket de test"))
    def imprimer_un_ticket_de_test(self, request, queryset):
        """
        Envoie un ticket de test sur les imprimantes selectionnees.
        / Sends a test ticket to the selected printers.

        LOCALISATION : Administration/admin/laboutik.py

        Sert a valider une imprimante qu'on vient de configurer, sans avoir a faire une
        vraie vente. L'envoi est SYNCHRONE : le gestionnaire doit voir l'erreur tout de
        suite si l'imprimante ne repond pas. C'est tout l'interet du test.
        / Synchronous on purpose: the manager must see the error immediately.
        """
        from laboutik.printing import imprimer_ticket_de_test

        for printer in queryset:
            resultat = imprimer_ticket_de_test(printer)

            if resultat["ok"]:
                self.message_user(request, _(
                    "%(nom)s : ticket de test envoyé."
                ) % {"nom": printer.name}, level=messages.SUCCESS)
            else:
                self.message_user(request, _(
                    "%(nom)s : échec — %(err)s"
                ) % {
                    "nom": printer.name,
                    "err": resultat.get("error", ""),
                }, level=messages.ERROR)

    fieldsets = (
        (_('General'), {
            'fields': (
                'name',
                'printer_type',
                'dots_per_line',
                'active',
            ),
        }),
        (_('Sunmi Cloud'), {
            'fields': (
                'sunmi_serial_number',
            ),
            'description': _(
                "Serial number required for Sunmi Cloud printers only."
            ),
        }),
        (_('Sunmi LAN'), {
            'fields': (
                'ip_address',
            ),
            'description': _(
                "IP address required for LAN printers only (same subnet)."
            ),
        }),
    )

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


class TerminalForm(forms.ModelForm):
    """
    Formulaire d'un terminal : creation et edition.
    / Terminal form: creation and editing.

    LOCALISATION : Administration/admin/laboutik.py

    A LA CREATION, le gestionnaire declare un appareil : son nom et ce qu'il sait faire.
    L'enregistrement fabrique un code PIN, qu'il ira taper sur l'appareil.

    A L'EDITION, il lui branche une imprimante. Le lecteur de carte bancaire, lui, se branche
    depuis SON propre ecran (« TPE bancaires ») : c'est le lecteur qu'on deplace d'un appareil
    a l'autre, le lien vit donc de son cote.

    LE ROLE « TIREUSE » N'EST PAS PROPOSE. Une tireuse se cree depuis son propre ecran
    (Tireuses) : elle porte du metier — un fut, un debitmetre, un prix — et c'est elle qui
    fabrique son terminal. Un terminal de role Tireuse cree ici n'aurait aucune tireuse
    derriere lui, et l'appairage echouerait.
    / The TAP role is NOT offered: a tap is created from its own screen and issues its own
    terminal. A tap-role terminal created here would have no tap behind it.
    """

    class Meta:
        model = Terminal
        # Ni le compte (term_user), ni le lecteur de carte ne sont dans ce formulaire :
        #
        # - term_user est le COMPTE de l'appareil. Il n'existe pas avant l'appairage, et il
        #   est pose par le claim. Le gestionnaire ne doit jamais le choisir a la main : le
        #   rattacher a un autre compte casserait le lien avec la cle d'API et la revocation.
        #   Il apparait en lecture seule dans TerminalAdmin.
        #
        # - le lecteur de carte (TPEBancaire) est un objet a part, qui designe LUI l'appareil
        #   sur lequel il est branche. On le branche depuis « TPE bancaires ».
        # / Neither the account (posed by the claim) nor the card reader (a separate object)
        # belongs in this form.
        fields = ["name", "terminal_role", "printer", "archived"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # Le role ne se choisit qu'A LA CREATION. Une fois l'appareil appaire, en changer
        # rendrait sa cle d'API incoherente avec ce qu'il est cense faire.
        #
        # ON TESTE _state.adding, PAS self.instance.pk.
        # Terminal.id est un UUIDField(default=uuid4) : son pk est donc DEJA REMPLI sur une
        # instance neuve, avant meme l'enregistrement. Tester `self.instance.pk` renverrait
        # toujours vrai, et le formulaire de creation se croirait en edition.
        # / Test _state.adding, NOT self.instance.pk: the UUID pk is already filled on a
        # brand-new instance, so testing pk would always be true.
        # / Same trap as AuthBillet.TermUser.save().
        terminal_deja_enregistre = (
            self.instance is not None and not self.instance._state.adding
        )
        if terminal_deja_enregistre:
            self.fields["terminal_role"].disabled = True
        else:
            from AuthBillet.models import TibilletUser

            self.fields["terminal_role"].choices = [
                (valeur, libelle)
                for valeur, libelle in TibilletUser.TERMINAL_ROLE_CHOICES
                if valeur != TibilletUser.ROLE_TIREUSE
            ]


# --- Lecteur de carte bancaire (TPE) ---
# / Card reader

class TPEBancaireForm(forms.ModelForm):
    """
    Formulaire d'un lecteur de carte bancaire.
    / Card reader form.

    LOCALISATION : Administration/admin/laboutik.py

    C'est ICI qu'on branche un lecteur sur un appareil, et qu'on l'en debranche pour le
    mettre ailleurs : le lien est porte par le lecteur, parce que c'est le lecteur qu'on
    deplace physiquement. Un seul geste, sur l'objet qu'on a en main.
    """

    class Meta:
        model = TPEBancaire
        fields = ["name", "tpe_type", "terminal", "registration_code", "active"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # On ne propose que les appareils en service.
        # / Only offer devices that are in service.
        self.fields["terminal"].queryset = Terminal.objects.filter(
            archived=False,
        ).order_by("name")
        self.fields["terminal"].empty_label = _("— Non branché —")

    def clean(self):
        cleaned_data = super().clean()

        deja_enregistre_chez_stripe = bool(
            self.instance.stripe_id if self.instance and not self.instance._state.adding
            else None
        )
        code_enregistrement = cleaned_data.get("registration_code")

        if not code_enregistrement and not deja_enregistre_chez_stripe:
            raise ValidationError({
                "registration_code": _(
                    "Le code d'enregistrement est obligatoire pour enregistrer le lecteur "
                    "chez Stripe."
                ),
            })

        # LE MEME LECTEUR PHYSIQUE NE PEUT PAS EXISTER DEUX FOIS.
        #
        # Sinon deux fiches piloteraient le meme appareil : un client verrait s'afficher, sur
        # le lecteur devant lui, le montant de la vente d'a cote — et pourrait la payer.
        # Stripe refuserait de toute facon un code deja consomme, mais son message est
        # cryptique : on le dit ici, clairement, avant meme de l'appeler.
        # / The same physical reader cannot exist twice.
        if code_enregistrement:
            un_autre_lecteur_a_deja_ce_code = TPEBancaire.objects.filter(
                registration_code=code_enregistrement,
            ).exclude(
                pk=self.instance.pk if self.instance else None,
            ).exists()

            if un_autre_lecteur_a_deja_ce_code:
                raise ValidationError({
                    "registration_code": _(
                        "Ce code est déjà utilisé par un autre lecteur. "
                        "Un lecteur physique ne peut être enregistré qu'une fois."
                    ),
                })

        return cleaned_data

    def _post_clean(self):
        # L'enregistrement chez Stripe se fait ICI, pas dans save_model : un code refuse par
        # Stripe doit invalider le formulaire — l'admin voit l'erreur sous le champ, et rien
        # n'est enregistre — au lieu de creer un lecteur orphelin, sans identifiant Stripe.
        # / Stripe registration happens HERE, not in save_model: a rejected code must
        # invalidate the form instead of creating an orphan reader with no Stripe id.
        super()._post_clean()

        if self.errors:
            return

        # Deja enregistre : rien a faire.
        # / Already registered: nothing to do.
        if self.instance.stripe_id:
            return

        try:
            self.instance.appairer_chez_stripe()
        except Exception as erreur_stripe:
            self.add_error("registration_code", _(
                "Échec de l'enregistrement du lecteur chez Stripe : %(err)s"
            ) % {"err": erreur_stripe})


@admin.register(TPEBancaire, site=staff_admin_site)
class TPEBancaireAdmin(ModelAdmin):
    """
    Admin des lecteurs de carte bancaire.
    / Card reader admin.

    LOCALISATION : Administration/admin/laboutik.py
    """
    compressed_fields = True
    warn_unsaved_form = True
    form = TPEBancaireForm

    list_display = ('name', 'tpe_type', 'terminal', 'etat_chez_stripe', 'active')
    list_select_related = ('terminal',)
    list_filter = ['tpe_type', 'active']
    search_fields = ['name', 'stripe_id']

    @admin.display(description=_("Stripe"))
    def etat_chez_stripe(self, tpe):
        """
        Le lecteur est-il enregistre chez Stripe ? Sans cela, il ne peut rien encaisser.
        / Is the reader registered at Stripe? Without it, it cannot take payments.
        """
        if tpe.est_appaire_chez_stripe():
            return format_html(
                '<span style="color: #16a34a; font-weight: 600;">✓ {}</span>',
                _("Enregistré"),
            )
        return format_html(
            '<span style="color: #dc2626;">✗ {}</span>',
            _("Non enregistré"),
        )

    actions = ['verifier_le_statut_chez_stripe']

    @admin.action(description=_("Vérifier le statut chez Stripe"))
    def verifier_le_statut_chez_stripe(self, request, queryset):
        """
        Demande a Stripe l'etat reel des lecteurs selectionnes (en ligne, hors ligne...).
        / Asks Stripe for the real state of the selected readers.

        LOCALISATION : Administration/admin/laboutik.py

        Utile a l'installation : on sait tout de suite si le lecteur est joignable, sans
        avoir a lancer une vraie vente.
        """
        for tpe in queryset:
            if not tpe.est_appaire_chez_stripe():
                self.message_user(request, _(
                    "%(nom)s : pas encore enregistré chez Stripe."
                ) % {"nom": tpe.name}, level=messages.WARNING)
                continue

            try:
                statut = tpe.statut_chez_stripe()
                self.message_user(request, _(
                    "%(nom)s : %(statut)s"
                ) % {"nom": tpe.name, "statut": statut}, level=messages.SUCCESS)
            except Exception as erreur_stripe:
                self.message_user(request, _(
                    "%(nom)s : échec — %(err)s"
                ) % {"nom": tpe.name, "err": erreur_stripe}, level=messages.ERROR)

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


@admin.register(Terminal, site=staff_admin_site)
class TerminalAdmin(ModelAdmin):
    """
    Admin des terminaux : tablettes Sunmi, Raspberry Pi, bornes libre-service.
    / Terminal admin: Sunmi tablets, Raspberry Pi, self-service kiosks.

    LOCALISATION : Administration/admin/laboutik.py
    """
    compressed_fields = True
    warn_unsaved_form = True
    form = TerminalForm

    list_display = (
        'name', 'terminal_role', 'etat_de_l_appairage',
        'printer', 'lecteur_de_carte', 'archived',
    )
    list_select_related = ('term_user', 'printer', 'tpe')
    list_filter = ['terminal_role', 'archived']
    search_fields = ['name']

    # Le compte de l'appareil est en LECTURE SEULE. Il est pose par l'appairage (le claim),
    # jamais choisi a la main : le rattacher a un autre compte casserait le lien avec sa
    # cle d'API et la revocation. On le montre, on ne le touche pas.
    # / The device's account is READ-ONLY: it is set by pairing, never chosen by hand.
    readonly_fields = ('compte_de_l_appareil',)

    @admin.display(description=_("Compte de l'appareil"))
    def compte_de_l_appareil(self, terminal):
        """
        Le compte pose par l'appairage. Vide tant que l'appareil n'a pas ete appaire.
        / The account set by pairing. Empty until the device has been paired.
        """
        if terminal.term_user is None:
            return _("— en attente d'appairage —")
        # first_name porte le nom lisible ; l'email est synthetique.
        # / first_name carries the readable name; the email is synthetic.
        return terminal.term_user.first_name or terminal.term_user.email

    @admin.display(description=_("TPE bancaire"))
    def lecteur_de_carte(self, terminal):
        """
        Le lecteur de carte branche sur cet appareil, s'il y en a un.
        / The card reader plugged into this device, if any.

        En lecture seule : on le branche et on le debranche depuis « TPE bancaires », parce
        que c'est le lecteur qu'on deplace, pas l'appareil.
        """
        lecteur = getattr(terminal, "tpe", None)
        if lecteur is None:
            return "—"
        return lecteur.name

    @admin.display(description=_("État"))
    def etat_de_l_appairage(self, terminal):
        """
        Ou en est cet appareil : appaire, en attente avec son code PIN, ou a relancer.
        / Where this device stands: paired, waiting with its PIN, or needing a fresh PIN.

        LOCALISATION : Administration/admin/laboutik.py

        C'est la colonne que le gestionnaire regarde : elle lui donne le code a taper sur
        l'appareil, sans qu'il ait a aller le chercher ailleurs.
        """
        if terminal.est_appaire():
            if terminal.term_user.is_active:
                return format_html(
                    '<span style="color: #16a34a; font-weight: 600;">✓ {}</span>',
                    _("Appairé"),
                )
            return format_html(
                '<span style="color: #dc2626; font-weight: 600;">✗ {}</span>',
                _("Révoqué"),
            )

        code_pin = terminal.code_pin_en_attente()
        if code_pin is None:
            return format_html(
                '<span style="color: #999;">{}</span>',
                _("Code PIN expiré — à régénérer"),
            )

        code_lisible = f"{str(code_pin)[:3]} {str(code_pin)[3:]}"
        return format_html(
            '<span style="font-family: monospace; font-size: 1.15em; font-weight: 700; '
            'letter-spacing: 0.1em;">{}</span>',
            code_lisible,
        )

    actions = ['revoquer_les_terminaux', 'generer_un_nouveau_code_pin']

    @admin.action(description=_("Générer un nouveau code PIN (appairer un autre appareil)"))
    def generer_un_nouveau_code_pin(self, request, queryset):
        """
        Redonne un code PIN au terminal, pour y appairer un appareil.
        / Issues a fresh PIN so a device can be paired onto this terminal.

        LOCALISATION : Administration/admin/laboutik.py

        C'est le geste a faire quand l'appareil est perdu, vole, ou grille. Le terminal, lui,
        SURVIT : il garde son imprimante, son lecteur de carte, et la tireuse qui le designe
        garde tout son historique. Seul le materiel est remplace.

        Trois situations, traitees dans cet ordre :

        1. Le terminal est appaire. On revoque d'abord l'appareil actuel — son compte ET sa
           cle — puis on detache le compte. INDISPENSABLE : sans revocation, l'appareil perdu
           continuerait de fonctionner avec sa cle, qui est stockee dessus.
        2. Un code PIN circule encore (il a juste expire). On le regenere.
        3. Le terminal n'a ni compte ni code. On lui en fabrique un.
        """
        from discovery.models import PairingDevice
        from discovery.services import fabriquer_le_code_pin_d_appairage

        RELATIONS_VERS_LES_CLES = ('laboutik_api_key', 'tireuse_api_key')
        nombre_de_codes_generes = 0

        for terminal in queryset:
            # 1. Couper l'acces de l'appareil actuel, s'il y en a un.
            # / 1. Cut off the current device's access, if any.
            if terminal.est_appaire():
                compte_de_l_ancien_appareil = terminal.term_user

                compte_de_l_ancien_appareil.is_active = False
                compte_de_l_ancien_appareil.save(update_fields=['is_active'])

                for nom_de_la_relation in RELATIONS_VERS_LES_CLES:
                    cle_api = getattr(compte_de_l_ancien_appareil, nom_de_la_relation, None)
                    if cle_api is not None:
                        cle_api.revoked = True
                        cle_api.save(update_fields=['revoked'])

                # On detache le compte sans le supprimer : il reste comme trace de l'appareil
                # remplace (sa derniere connexion, notamment).
                # / Detach without deleting: it stays as a trace of the replaced device.
                terminal.term_user = None
                terminal.save(update_fields=['term_user'])

            # 2. Un code circule encore ? On le regenere plutot que d'en empiler un second.
            # / 2. A PIN still circulating? Regenerate it rather than stacking a second one.
            code_en_circulation = PairingDevice.objects.filter(
                cible_uuid=terminal.id,
                claimed_at__isnull=True,
            ).first()

            if code_en_circulation is not None:
                code_en_circulation.regenerer_le_pin()
            else:
                # 3. Rien en circulation : on en fabrique un.
                # / 3. Nothing circulating: issue one.
                fabriquer_le_code_pin_d_appairage(terminal)

            nombre_de_codes_generes += 1

        self.message_user(request, _(
            "%(nb)s code(s) PIN généré(s). Tapez-le sur l'appareil pour l'appairer."
        ) % {"nb": nombre_de_codes_generes})

    @admin.action(description=_("Révoquer le terminal (coupe son accès)"))
    def revoquer_les_terminaux(self, request, queryset):
        """
        Coupe l'acces d'un terminal. Il faut agir sur DEUX leviers.
        / Cuts a terminal's access. TWO levers are needed.

        LOCALISATION : Administration/admin/laboutik.py

        1. term_user.is_active = False
           Coupe le pont d'authentification et refuse les reconnexions WebSocket.

        2. La cle d'API : revoked = True
           Coupe l'en-tete Api-Key. INDISPENSABLE en plus du point 1 : la cle est stockee
           sur l'appareil. Sans la revoquer, il suffirait de reactiver le compte pour que
           l'appareil se reconnecte tout seul.

        Un terminal porte UNE cle, mais de deux classes possibles : LaBoutikAPIKey pour une
        caisse ou une borne, TireuseAPIKey pour une tireuse. Les permissions de controlvanne
        s'appuient sur une classe distincte : on coupe donc les deux, sans se demander de
        quel role il s'agit.
        / A terminal holds ONE key, of one of two classes. We revoke both, regardless of role.
        """
        nombre_de_terminaux_revoques = 0

        # Le nom de la relation inverse pour chaque classe de cle.
        # / The reverse relation name for each key class.
        RELATIONS_VERS_LES_CLES = ('laboutik_api_key', 'tireuse_api_key')

        for terminal in queryset:
            if not terminal.term_user:
                continue

            terminal.term_user.is_active = False
            terminal.term_user.save(update_fields=['is_active'])

            for nom_de_la_relation in RELATIONS_VERS_LES_CLES:
                # La cle peut ne pas exister : le champ user est nullable des deux cotes
                # (cles creees avant le pont V2), et un terminal n'a de toute facon qu'une
                # seule des deux classes.
                # / The key may not exist: the user field is nullable on both classes.
                cle_api = getattr(terminal.term_user, nom_de_la_relation, None)
                if cle_api is not None:
                    cle_api.revoked = True
                    cle_api.save(update_fields=['revoked'])

            nombre_de_terminaux_revoques += 1

        self.message_user(request, _(
            "%(nb)s terminal(aux) révoqué(s)."
        ) % {"nb": nombre_de_terminaux_revoques})

    def save_model(self, request, obj, form, change):
        """
        A la creation d'un terminal, on lui fabrique aussitot son code PIN.
        / On terminal creation, issue its PIN right away.

        LOCALISATION : Administration/admin/laboutik.py

        C'est ce qui rend l'appareil appairable : le gestionnaire declare l'appareil ici,
        lit le code dans la colonne « Etat », et va le taper dessus.

        Le code est fabrique par un appel EXPLICITE, pas par un signal : un signal sur
        Terminal en fabriquerait un a chaque creation, y compris dans les tests qui creent
        des terminaux directement. Voir discovery/services.py.
        """
        super().save_model(request, obj, form, change)

        terminal_vient_d_etre_cree = not change
        if terminal_vient_d_etre_cree:
            from discovery.services import fabriquer_le_code_pin_d_appairage

            fabriquer_le_code_pin_d_appairage(obj)

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


@admin.register(CartePrimaire, site=staff_admin_site)
class CartePrimaireAdmin(ModelAdmin):
    """Admin pour les cartes primaires (operateurs de caisse).
    Admin for primary cards (POS operators).
    LOCALISATION : Administration/admin/laboutik.py"""
    compressed_fields = True
    warn_unsaved_form = True

    list_display = ('carte', 'edit_mode', 'datetime')
    list_filter = ['edit_mode']
    search_fields = ['carte__tag_id', 'carte__number']
    filter_horizontal = ('points_de_vente',)

    # Champ de recherche (select2) au lieu d'une liste deroulante de toutes les
    # cartes. On tape le numero imprime (ou le tag NFC) et les resultats
    # arrivent en direct. La recherche s'appuie sur les search_fields de
    # QrcodeCashless.admin.CarteCashlessAdmin (tag_id, number, user__email).
    # / Search field (select2) instead of a dropdown listing every card. Type the
    # printed number (or the NFC tag) and results stream in. The search relies on
    # CarteCashlessAdmin.search_fields.
    #
    # Bonus securite : la vue d'autocompletion de Django passe par
    # CarteCashlessAdmin.get_queryset(), qui filtre deja sur le tenant courant.
    # / Security bonus: Django's autocomplete view goes through
    # CarteCashlessAdmin.get_queryset(), already filtered on the current tenant.
    autocomplete_fields = ('carte',)

    fieldsets = (
        (None, {
            'fields': (
                'carte',
                'edit_mode',
                'points_de_vente',
            ),
        }),
    )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """
        Restreint les cartes selectionnables a celles du lieu courant.
        / Restricts selectable cards to the current venue's.

        INDISPENSABLE meme avec autocomplete_fields : l'autocompletion ne pilote
        que l'affichage et la recherche. C'est le queryset du champ qui VALIDE la
        valeur postee. Sans ce filtre, un pk forge pointant vers la carte d'un
        autre lieu serait accepte — CarteCashless est en SHARED_APPS (schema
        public), il n'y a aucune isolation automatique.
        / REQUIRED even with autocomplete_fields: autocompletion only drives
        display and search. The field's queryset is what VALIDATES the posted
        value. Without this filter, a forged pk pointing at another venue's card
        would be accepted — CarteCashless lives in SHARED_APPS (public schema).
        """
        if db_field.name == 'carte':
            kwargs['queryset'] = CarteCashless.objects.filter(
                detail__origine=connection.tenant,
            ).select_related('detail')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


@admin.register(CategorieTable, site=staff_admin_site)
class CategorieTableAdmin(ModelAdmin):
    """Admin minimal pour les categories de table (Phase 4 = restaurant).
    Minimal admin for table categories (Phase 4 = restaurant).
    LOCALISATION : Administration/admin/laboutik.py"""
    list_display = ('name', 'icon')
    search_fields = ['name']

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


@admin.register(Table, site=staff_admin_site)
class TableAdmin(ModelAdmin):
    """Admin minimal pour les tables de restaurant (Phase 4 = restaurant).
    Minimal admin for restaurant tables (Phase 4 = restaurant).
    LOCALISATION : Administration/admin/laboutik.py"""
    list_display = ('name', 'categorie', 'statut', 'ephemere', 'archive')
    list_filter = ['statut', 'categorie', 'archive']
    search_fields = ['name']
    ordering = ('poids', 'name')

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


# --- Commandes de restaurant (Phase 4) ---
# --- Restaurant orders (Phase 4) ---

class ArticleCommandeSauvegardeInline(TabularInline):
    """Inline lecture seule pour les articles d'une commande.
    Read-only inline for order articles.
    LOCALISATION : Administration/admin/laboutik.py"""
    model = ArticleCommandeSauvegarde
    extra = 0
    fields = ('product', 'price', 'qty', 'reste_a_payer', 'reste_a_servir', 'statut')
    readonly_fields = ('product', 'price', 'qty', 'reste_a_payer', 'reste_a_servir', 'statut')

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(CommandeSauvegarde, site=staff_admin_site)
class CommandeSauvegardeAdmin(ModelAdmin):
    """Admin lecture seule pour l'historique des commandes de restaurant.
    Read-only admin for restaurant order history.
    LOCALISATION : Administration/admin/laboutik.py"""
    list_display = ('uuid', 'table', 'statut', 'responsable', 'datetime', 'archive')
    list_filter = ['statut', 'archive']
    search_fields = ['uuid', 'commentaire']
    ordering = ('-datetime',)
    readonly_fields = (
        'uuid', 'service', 'responsable', 'table', 'datetime',
        'statut', 'commentaire', 'archive',
    )
    inlines = [ArticleCommandeSauvegardeInline]

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


# --- Cloture de caisse (Phase 5) ---
# --- Cash register closure (Phase 5) ---

# --- Helpers d'export pour ClotureCaisseAdmin ---
# Definis HORS de la classe pour eviter qu'Unfold les wrappe avec @action.
# / Defined OUTSIDE the class to prevent Unfold from wrapping them with @action.

def _euros(centimes):
    """Convertit des centimes (int) en euros (float arrondi). / Converts cents to euros."""
    if centimes is None:
        return 0.0
    return round(centimes / 100, 2)


def _ecrire_rapport_csv_excel(writer, cloture, rapport):
    """
    Ecrit le rapport section par section dans un writer generique.
    Le writer doit implementer append_title, append_row, append_header, append_blank.
    / Writes the report section by section into a generic writer.
    """
    e = _euros

    # --- En-tete ---
    writer.append_title(str(_("Closure report")))
    writer.append_row([str(_("Level")), cloture.get_niveau_display()])
    writer.append_row([str(_("Number")), f"#{cloture.numero_sequentiel}"])
    writer.append_row([str(_("Point of sale")), cloture.point_de_vente.name if cloture.point_de_vente else "—"])
    writer.append_row([str(_("Responsible")), str(cloture.responsable or "—")])
    writer.append_row([str(_("Period")), f"{cloture.datetime_ouverture} → {cloture.datetime_cloture}"])
    writer.append_blank()

    # --- Section 1 : Totaux par moyen de paiement ---
    section = rapport.get("totaux_par_moyen", {})
    if section:
        writer.append_title(str(_("Totals by payment method")))
        writer.append_header([str(_("Payment method")), str(_("Amount"))])
        writer.append_row([str(_("Cash")), e(section.get("especes", 0))])
        writer.append_row([str(_("Credit card")), e(section.get("carte_bancaire", 0))])
        writer.append_row([str(_("Cashless")), e(section.get("cashless", 0))])
        for asset in section.get("cashless_detail", []):
            writer.append_row([f"  ↳ {asset['nom']} ({asset['code']})", e(asset["montant"])])
        writer.append_row([str(_("Check")), e(section.get("cheque", 0))])
        writer.append_row([str(_("Total")), e(section.get("total", 0))])
        writer.append_blank()

    # --- Section 2 : Detail des ventes ---
    section = rapport.get("detail_ventes", {})
    if section:
        writer.append_title(str(_("Sales detail")))
        writer.append_header([
            str(_("Category")), str(_("Product")),
            str(_("Sold")), str(_("Free")), str(_("Total qty")),
            str(_("HT")), str(_("VAT")), str(_("TTC")),
            str(_("Cost")), str(_("Profit")),
        ])
        for cat_nom, cat_data in section.items():
            for article in cat_data.get("articles", []):
                writer.append_row([
                    cat_nom, article.get("nom", "—"),
                    article.get("qty_vendus", 0), article.get("qty_offerts", 0),
                    article.get("qty_total", 0),
                    e(article.get("total_ht", 0)), e(article.get("total_tva", 0)),
                    e(article.get("total_ttc", 0)),
                    e(article.get("cout_total", 0)), e(article.get("benefice", 0)),
                ])
            writer.append_row([f"Total {cat_nom}", "", "", "", "", "", "", e(cat_data.get("total_ttc", 0)), "", ""])
        writer.append_blank()

    # --- Section 3 : TVA ---
    section = rapport.get("tva", {})
    if section:
        writer.append_title(str(_("VAT breakdown")))
        writer.append_header([str(_("Rate")), str(_("HT")), str(_("VAT")), str(_("TTC"))])
        for taux, data in section.items():
            writer.append_row([taux, e(data.get("total_ht", 0)), e(data.get("total_tva", 0)), e(data.get("total_ttc", 0))])
        writer.append_blank()

    # --- Section 4 : Solde caisse ---
    section = rapport.get("solde_caisse", {})
    if section:
        writer.append_title(str(_("Cash register balance")))
        writer.append_row([str(_("Opening float")), e(section.get("fond_de_caisse", 0))])
        writer.append_row([str(_("Cash income")), e(section.get("entrees_especes", 0))])
        writer.append_row([str(_("Balance")), e(section.get("solde", 0))])
        writer.append_blank()

    # --- Section 5 : Recharges ---
    section = rapport.get("recharges", {})
    if section and section.get("detail"):
        writer.append_title(str(_("Cashless top-ups")))
        writer.append_header([str(_("Product")), str(_("Currency")), str(_("Payment method")), str(_("Amount")), str(_("Count"))])
        for cle, rec in section["detail"].items():
            writer.append_row([rec.get("nom_produit", "—"), rec.get("nom_monnaie", "—"), rec.get("moyen_paiement", "—"), e(rec.get("total", 0)), rec.get("nb", 0)])
        writer.append_row([str(_("Total")), "", "", e(section.get("total", 0)), ""])
        writer.append_blank()

    # --- Section 6 : Adhesions ---
    section = rapport.get("adhesions", {})
    if section and section.get("detail"):
        writer.append_title(str(_("Memberships")))
        writer.append_header([str(_("Product")), str(_("Price tier")), str(_("Payment method")), str(_("Count")), str(_("Amount"))])
        for cle, adh in section["detail"].items():
            writer.append_row([adh.get("nom_produit", "—"), adh.get("nom_tarif", "—"), adh.get("moyen_paiement", "—"), adh.get("nb", 0), e(adh.get("total", 0))])
        writer.append_row([str(_("Total")), "", "", section.get("nb", 0), e(section.get("total", 0))])
        writer.append_blank()

    # --- Section 7 : Remboursements ---
    section = rapport.get("remboursements", {})
    if section:
        writer.append_title(str(_("Refunds")))
        writer.append_row([str(_("Count")), section.get("nb", 0)])
        writer.append_row([str(_("Total")), e(section.get("total", 0))])
        writer.append_blank()

    # --- Section 8 : Habitus ---
    section = rapport.get("habitus", {})
    if section:
        writer.append_title(str(_("Customer statistics")))
        writer.append_row([str(_("Cards used")), section.get("nb_cartes", 0)])
        writer.append_row([str(_("Total spent")), e(section.get("total", 0))])
        writer.append_row([str(_("Average basket")), e(section.get("panier_moyen", 0))])
        writer.append_row([str(_("Median spend")), e(section.get("depense_mediane", 0))])
        writer.append_row([str(_("Median top-up")), e(section.get("recharge_mediane", 0))])
        writer.append_blank()

    # --- Section 9 : Billets ---
    section = rapport.get("billets", {})
    if section and section.get("detail"):
        writer.append_title(str(_("Tickets")))
        writer.append_header([str(_("Event")), str(_("Date")), str(_("Product / Price tier")), str(_("Count")), str(_("Amount"))])
        for cle, b in section["detail"].items():
            tarif_label = b.get("nom_produit", "")
            if b.get("nom_tarif"):
                tarif_label += f" / {b['nom_tarif']}"
            writer.append_row([b.get("nom_event", "—"), b.get("date_event", "—"), tarif_label, b.get("nb", 0), e(b.get("total", 0))])
        writer.append_row([str(_("Total")), "", "", section.get("nb", 0), e(section.get("total", 0))])
        writer.append_blank()

    # --- Section 10 : Synthese ---
    section = rapport.get("synthese_operations", {})
    if section:
        writer.append_title(str(_("Operations summary")))
        writer.append_header([str(_("Operation")), str(_("Cash")), str(_("Credit card")), str(_("Cashless")), str(_("Total"))])
        for op_nom, op_data in section.items():
            writer.append_row([op_nom.title(), e(op_data.get("especes", 0)), e(op_data.get("carte_bancaire", 0)), e(op_data.get("cashless", 0)), e(op_data.get("total", 0))])
        writer.append_blank()


@admin.register(ClotureCaisse, site=staff_admin_site)
class ClotureCaisseAdmin(ModelAdmin):
    """Admin lecture seule pour les clotures de caisse.
    Document comptable immuable — aucune modification possible.
    Read-only admin for cash register closures.
    Immutable accounting document — no modification allowed.
    LOCALISATION : Administration/admin/laboutik.py"""

    def _recalculer_rapport(self, cloture):
        """
        Recalcule le rapport a la volee depuis les LigneArticle.
        Garantit que les modifications de reports.py sont visibles
        sans avoir a regenerer les clotures existantes.
        / Recalculates the report on-the-fly from LigneArticle.
        Ensures reports.py changes are visible
        without regenerating existing closures.
        """
        from laboutik.reports import RapportComptableService
        service = RapportComptableService(
            point_de_vente=cloture.point_de_vente,
            datetime_debut=cloture.datetime_ouverture,
            datetime_fin=cloture.datetime_cloture,
        )
        return service.generer_rapport_complet()

    list_display = (
        'datetime_cloture',
        'niveau', 'numero_sequentiel',
        'responsable',
        'ca_ttc_euros',
    )
    list_display_links = ('datetime_cloture',)
    list_filter = ['niveau']
    search_fields = ['point_de_vente__name', 'responsable__email']
    ordering = ('-datetime_cloture',)
    # Pas de fieldsets — tout le contenu est dans le change_form_before_template.
    # Le rapport comptable remplace le formulaire standard.
    # / No fieldsets — all content is in the change_form_before_template.
    # The accounting report replaces the standard form.
    fieldsets = ()
    readonly_fields = ()
    change_form_before_template = "admin/cloture/rapport_before.html"

    @admin.display(description=_("Revenue incl. tax"))
    def ca_ttc_euros(self, obj):
        """Affiche le total general en euros. / Displays grand total in euros."""
        from django.utils.html import format_html
        euros = obj.total_general / 100
        euros_formate = f"{euros:,.2f} €".replace(",", " ")
        return format_html('<span style="font-variant-numeric: tabular-nums;">{}</span>', euros_formate)

    def changelist_view(self, request, extra_context=None):
        """
        Injecte l'URL de l'export fiscal dans le contexte du changelist.
        Affiche un bandeau avec le bouton "Export fiscal" en haut de la liste.
        / Injects the fiscal export URL into the changelist context.
        Displays a banner with the "Export fiscal" button at the top of the list.
        LOCALISATION : Administration/admin/laboutik.py
        """
        extra_context = extra_context or {}
        extra_context['export_fiscal_url'] = '/laboutik/caisse/export-fiscal/'
        extra_context['export_fec_url'] = '/laboutik/caisse/export-fec/'
        extra_context['export_csv_comptable_url'] = '/laboutik/caisse/export-csv-comptable/'
        extra_context['rapport_temps_reel_url'] = '/laboutik/caisse/rapport-temps-reel/'
        return super().changelist_view(request, extra_context)

    list_before_template = "admin/cloture/changelist_before.html"

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        """
        Injecte le rapport recalcule dans le contexte du template before.
        Le template before affiche le rapport complet + boutons export.
        / Injects the recalculated report into the before template context.
        The before template displays the full report + export buttons.
        LOCALISATION : Administration/admin/laboutik.py
        """
        extra_context = extra_context or {}
        if object_id:
            cloture = get_object_or_404(ClotureCaisse, pk=object_id)
            extra_context["rapport"] = self._recalculer_rapport(cloture)
            extra_context["cloture_obj"] = cloture
        return super().changeform_view(request, object_id, form_url, extra_context)

    def get_urls(self):
        """
        URLs custom pour les exports (CSV, PDF, Excel, fiscal ZIP).
        Le rapport est affiche via changeform_view (pas d'URL custom).
        / Custom URLs for exports (CSV, PDF, Excel, fiscal ZIP).
        The report is displayed via changeform_view (no custom URL).
        """
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path(
                '<path:object_id>/exporter-csv/',
                self.admin_site.admin_view(self.exporter_csv),
                name='laboutik_cloturecaisse_exporter_csv',
            ),
            path(
                '<path:object_id>/exporter-pdf/',
                self.admin_site.admin_view(self.exporter_pdf),
                name='laboutik_cloturecaisse_exporter_pdf',
            ),
            path(
                '<path:object_id>/exporter-excel/',
                self.admin_site.admin_view(self.exporter_excel),
                name='laboutik_cloturecaisse_exporter_excel',
            ),
            path(
                '<path:object_id>/exporter-fec/',
                self.admin_site.admin_view(self.exporter_fec),
                name='laboutik_cloturecaisse_exporter_fec',
            ),
            path(
                '<path:object_id>/exporter-csv-comptable/',
                self.admin_site.admin_view(self.exporter_csv_comptable),
                name='laboutik_cloturecaisse_exporter_csv_comptable',
            ),
        ]
        return custom_urls + urls

    def exporter_csv(self, request, object_id):
        """
        Exporte le rapport de cloture en CSV structure (delimiteur ;).
        Meme structure que le HTML, pas de JSON brut.
        / Exports the closure report as structured CSV (delimiter ;).
        Same structure as HTML, no raw JSON.
        LOCALISATION : Administration/admin/laboutik.py
        """
        import csv
        from django.http import HttpResponse

        cloture = get_object_or_404(ClotureCaisse, pk=object_id)
        rapport = self._recalculer_rapport(cloture)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"rapport_{cloture.get_niveau_display()}_{cloture.numero_sequentiel}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')

        csv_writer = csv.writer(response, delimiter=';')

        # Adaptateur CSV pour le writer generique
        # / CSV adapter for the generic writer
        class CsvWriterAdapter:
            def append_title(self, titre):
                csv_writer.writerow([])
                csv_writer.writerow([titre.upper()])
            def append_header(self, cols):
                csv_writer.writerow(cols)
            def append_row(self, cols):
                csv_writer.writerow(cols)
            def append_blank(self):
                csv_writer.writerow([])

        _ecrire_rapport_csv_excel(CsvWriterAdapter(), cloture, rapport)
        return response

    def exporter_pdf(self, request, object_id):
        """
        Exporte le rapport de cloture en PDF A4 (WeasyPrint).
        / Exports the closure report as A4 PDF (WeasyPrint).
        LOCALISATION : Administration/admin/laboutik.py
        """
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        from weasyprint import HTML
        from BaseBillet.models import Configuration

        cloture = get_object_or_404(ClotureCaisse, pk=object_id)
        rapport = self._recalculer_rapport(cloture)
        config = Configuration.get_solo()

        # Adresse complete assemblee depuis les parties disponibles
        # / Full address assembled from available parts
        parties_adresse = []
        if config.adress:
            parties_adresse.append(config.adress)
        if config.postal_code:
            parties_adresse.append(str(config.postal_code))
        if config.city:
            parties_adresse.append(config.city)

        context = {
            "cloture": cloture,
            "rapport": rapport,
            "config_org": config.organisation or "",
            "config_siret": config.siren or "",
            "config_address": " ".join(parties_adresse),
            "now": timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M"),
        }

        html_string = render_to_string(
            "laboutik/pdf/rapport_comptable.html", context,
        )
        pdf_bytes = HTML(string=html_string).write_pdf()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = f"rapport_{cloture.get_niveau_display()}_{cloture.numero_sequentiel}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def exporter_excel(self, request, object_id):
        """
        Exporte le rapport de cloture en Excel (1 seul onglet, mise en forme soignee).
        Meme structure que le HTML, pas de JSON brut.
        / Exports the closure report as Excel (single sheet, clean formatting).
        Same structure as HTML, no raw JSON.
        LOCALISATION : Administration/admin/laboutik.py
        """
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        cloture = get_object_or_404(ClotureCaisse, pk=object_id)
        rapport = self._recalculer_rapport(cloture)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"

        # Styles
        titre_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11, color="FFFFFF")
        section_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        total_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        # Adaptateur Excel pour le writer generique
        # / Excel adapter for the generic writer
        current_row = [1]
        max_cols_used = [1]

        class ExcelWriterAdapter:
            def append_title(self, titre):
                row = current_row[0]
                cell = ws.cell(row=row, column=1, value=titre)
                cell.font = section_font
                cell.fill = section_fill
                # Etendre le fond sur 10 colonnes pour l'effet visuel
                # / Extend background across 10 columns for visual effect
                for col in range(2, 11):
                    c = ws.cell(row=row, column=col)
                    c.fill = section_fill
                current_row[0] += 1

            def append_header(self, cols):
                row = current_row[0]
                for idx, col_val in enumerate(cols, 1):
                    cell = ws.cell(row=row, column=idx, value=col_val)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                max_cols_used[0] = max(max_cols_used[0], len(cols))
                current_row[0] += 1

            def append_row(self, cols):
                row = current_row[0]
                for idx, col_val in enumerate(cols, 1):
                    cell = ws.cell(row=row, column=idx, value=col_val)
                    cell.border = thin_border
                    # Aligner les nombres a droite
                    # / Right-align numbers
                    if isinstance(col_val, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                        cell.number_format = '#,##0.00'
                max_cols_used[0] = max(max_cols_used[0], len(cols))
                current_row[0] += 1

            def append_blank(self):
                current_row[0] += 1

        _ecrire_rapport_csv_excel(ExcelWriterAdapter(), cloture, rapport)

        # Auto-largeur des colonnes / Auto-width columns
        for col_idx in range(1, max_cols_used[0] + 1):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"rapport_{cloture.get_niveau_display()}_{cloture.numero_sequentiel}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def exporter_fec(self, request, object_id):
        """
        Exporte une seule cloture au format FEC (18 colonnes).
        Le fichier contient 1 ecriture comptable equilibree (debits = credits).
        / Exports a single closure as FEC format (18 columns).
        The file contains 1 balanced accounting entry (debits = credits).
        LOCALISATION : Administration/admin/laboutik.py
        """
        from django.db import connection
        from django.http import HttpResponse

        from laboutik.fec import generer_fec

        cloture = get_object_or_404(ClotureCaisse, pk=object_id)

        # Generer le FEC pour cette seule cloture
        # / Generate FEC for this single closure
        schema = connection.schema_name
        contenu_bytes, nom_fichier, avertissements = generer_fec(
            ClotureCaisse.objects.filter(pk=object_id),
            schema,
        )

        response = HttpResponse(contenu_bytes, content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        return response

    def exporter_csv_comptable(self, request, object_id):
        """
        GET : mini-formulaire avec choix du profil comptable.
        POST : telecharge le CSV comptable pour cette cloture.
        / GET: mini form with accounting profile choice.
        POST: downloads the accounting CSV for this closure.
        LOCALISATION : Administration/admin/laboutik.py
        """
        from django.db import connection
        from django.http import HttpResponse

        from laboutik.csv_comptable import generer_csv_comptable
        from laboutik.profils_csv import PROFILS

        cloture = get_object_or_404(ClotureCaisse, pk=object_id)

        if request.method == "GET":
            # Renvoie un partial HTMX charge dans #detail-export-zone
            # Le formulaire POST pointe vers la meme URL admin
            # / Returns an HTMX partial loaded into #detail-export-zone
            from django.http import HttpResponse as HR
            from django.middleware.csrf import get_token
            profil_options = ''.join(
                f'<option value="{cle}">{p["nom"]}</option>'
                for cle, p in PROFILS.items()
            )
            csrf = get_token(request)
            html = f'''
            <div style="margin: 16px 0; padding: 16px; background: #f8f9fa; border-radius: 8px;
                        animation: fadeSlideIn 300ms ease both;" data-testid="detail-csv-comptable-form">
                <style>@keyframes fadeSlideIn {{ from {{ opacity: 0; transform: translateY(-6px); }} }}</style>
                <form method="post" style="display: flex; align-items: center; gap: 12px; flex-wrap: wrap;">
                    <input type="hidden" name="csrfmiddlewaretoken" value="{csrf}">
                    <label style="font-weight: 500; font-size: 0.85em; color: #333;">Profil :</label>
                    <select name="profil" style="padding: 6px 12px; border-radius: 4px; border: 1px solid #ccc; font-size: 0.85em;">
                        {profil_options}
                    </select>
                    <button type="submit" style="padding: 6px 16px; background: #8B5CF6; color: white;
                            border: none; border-radius: 6px; font-size: 0.85em; font-weight: 500; cursor: pointer;">
                        <i class="fas fa-download" aria-hidden="true"></i> Telecharger
                    </button>
                    <button type="button" onclick="document.getElementById(\'detail-export-zone\').innerHTML=\'\'"
                            style="padding: 6px 12px; background: none; border: none; color: #666;
                                   font-size: 0.85em; cursor: pointer;">
                        Annuler
                    </button>
                </form>
            </div>'''
            return HR(html)

        # --- POST : generation du CSV comptable ---
        # --- POST: accounting CSV generation ---
        profil_nom = request.POST.get('profil', '').strip()
        if profil_nom not in PROFILS:
            from django.http import HttpResponseBadRequest
            return HttpResponseBadRequest("Profil inconnu.")

        schema = connection.schema_name
        contenu_bytes, nom_fichier, avertissements = generer_csv_comptable(
            ClotureCaisse.objects.filter(pk=object_id),
            profil_nom,
            schema,
        )

        profil = PROFILS[profil_nom]
        content_type = 'text/csv; charset=' + profil["encodage"]
        response = HttpResponse(contenu_bytes, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        return response

    @admin.display(description=_("Integrity"))
    def badge_integrite(self, obj):
        """
        Badge vert si hash_lignes est present, tiret si vide.
        Pour les clotures M/A, le hash n'est pas applicable.
        / Green badge if hash_lignes present, dash if empty.
        For M/A closures, hash is not applicable.
        """
        from django.utils.html import format_html
        if obj.niveau != ClotureCaisse.JOURNALIERE:
            return format_html('<span style="color: gray;">—</span>')
        if obj.hash_lignes:
            return format_html('<span style="color: green;">✓</span>')
        return format_html('<span style="color: orange;">—</span>')

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


# --- Journal des impressions (conformite LNE exigence 9) ---
# --- Print audit log (LNE compliance req. 9) ---

@admin.register(ImpressionLog, site=staff_admin_site)
class ImpressionLogAdmin(ModelAdmin):
    """Admin lecture seule pour la tracabilite des impressions.
    Read-only admin for print tracking.
    LOCALISATION : Administration/admin/laboutik.py"""
    list_display = (
        'datetime', 'type_justificatif', 'is_duplicata',
        'uuid_transaction', 'printer', 'operateur', 'format_emission',
    )
    list_filter = ('type_justificatif', 'is_duplicata', 'format_emission')
    search_fields = ('uuid_transaction',)
    ordering = ('-datetime',)
    readonly_fields = (
        'uuid', 'datetime', 'ligne_article', 'uuid_transaction',
        'cloture', 'operateur', 'printer', 'type_justificatif',
        'is_duplicata', 'format_emission',
    )

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


# --- Journal des operations techniques (conformite LNE) ---
# --- Technical operations log (LNE compliance) ---

@admin.register(JournalOperation, site=staff_admin_site)
class JournalOperationAdmin(ModelAdmin):
    """Admin lecture seule pour le journal des operations techniques.
    Read-only admin for the technical operations log.
    LOCALISATION : Administration/admin/laboutik.py"""
    list_display = ('datetime', 'type_operation', 'operateur')
    list_filter = ('type_operation',)
    search_fields = ('operateur__email',)
    ordering = ('-datetime',)
    readonly_fields = ('uuid', 'type_operation', 'datetime', 'operateur', 'details', 'hmac_hash')

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


# --- Historique du fond de caisse ---
# --- Cash float history ---

# Helpers de formatage definis HORS de la classe pour eviter qu'Unfold les wrappe avec @action.
# / Formatting helpers defined OUTSIDE the class to prevent Unfold from wrapping them with @action.

def _euros_ancien(obj):
    """Formate l'ancien montant en euros. / Formats previous amount in euros."""
    return f"{obj.ancien_montant / 100:.2f} €"

_euros_ancien.short_description = _("Previous amount")


def _euros_nouveau(obj):
    """Formate le nouveau montant en euros. / Formats new amount in euros."""
    return f"{obj.nouveau_montant / 100:.2f} €"

_euros_nouveau.short_description = _("New amount")


@admin.register(HistoriqueFondDeCaisse, site=staff_admin_site)
class HistoriqueFondDeCaisseAdmin(ModelAdmin):
    """Admin lecture seule pour l'historique du fond de caisse.
    Read-only admin for cash float history.
    LOCALISATION : Administration/admin/laboutik.py"""
    list_display = ('datetime', _euros_ancien, _euros_nouveau, 'operateur')
    list_filter = ('point_de_vente',)
    search_fields = ('operateur__email',)
    ordering = ('-datetime',)
    readonly_fields = ('uuid', 'point_de_vente', 'operateur', 'datetime', 'ancien_montant', 'nouveau_montant', 'raison')

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


# --- Plan comptable (export comptable) ---
# --- Chart of accounts (accounting export) ---

@admin.register(CompteComptable, site=staff_admin_site)
class CompteComptableAdmin(ModelAdmin):
    """Admin CRUD pour les comptes du Plan Comptable General (PCG).
    Bandeau « Charger un plan comptable » en haut de la liste.
    CRUD admin for the French Chart of Accounts (PCG) entries.
    "Load a chart of accounts" banner at the top of the list.
    LOCALISATION : Administration/admin/laboutik.py"""
    compressed_fields = True
    warn_unsaved_form = True

    list_display = ('numero_de_compte', 'libelle_du_compte', 'nature_du_compte', 'taux_de_tva', 'est_actif')
    list_filter = ('nature_du_compte', 'est_actif')
    search_fields = ('numero_de_compte', 'libelle_du_compte')
    ordering = ('numero_de_compte',)

    list_before_template = "admin/comptable/changelist_before.html"

    def changelist_view(self, request, extra_context=None):
        """Injecte l'URL pour charger un plan comptable par defaut.
        / Injects the URL to load a default chart of accounts."""
        extra_context = extra_context or {}
        extra_context['charger_plan_url'] = '/laboutik/caisse/charger-plan-comptable/'
        return super().changelist_view(request, extra_context)

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)


# Helper module-level : libelle humain du moyen de paiement pour la liste admin
# Unfold wrappe les methodes du ModelAdmin — definir les helpers HORS de la classe.
# / Module-level helper: human label for payment method in admin list
def _display_moyen_paiement(obj):
    """Affiche le libelle humain du moyen de paiement (PaymentMethod.label).
    / Displays the human label of the payment method."""
    from BaseBillet.models import PaymentMethod
    # Chercher le label dans PaymentMethod.choices
    # / Look up the label in PaymentMethod.choices
    dict_labels = dict(PaymentMethod.choices)
    code = obj.moyen_de_paiement
    label = dict_labels.get(code, code)
    return f"{label} ({code})"
_display_moyen_paiement.short_description = _("Payment method")


@admin.register(MappingMoyenDePaiement, site=staff_admin_site)
class MappingMoyenDePaiementAdmin(ModelAdmin):
    """Admin CRUD pour le mapping moyen de paiement → compte de tresorerie.
    CRUD admin for payment method → treasury account mapping.
    LOCALISATION : Administration/admin/laboutik.py"""
    compressed_fields = True
    warn_unsaved_form = True

    list_display = (_display_moyen_paiement, 'libelle_moyen', 'compte_de_tresorerie')
    autocomplete_fields = ['compte_de_tresorerie']

    def get_form(self, request, obj=None, **kwargs):
        """
        Remplace le champ texte libre 'moyen_de_paiement' par un menu deroulant
        avec les choix PaymentMethod (libelles humains).
        / Replaces the free text 'moyen_de_paiement' field with a dropdown
        using PaymentMethod choices (human labels).
        LOCALISATION : Administration/admin/laboutik.py
        """
        form = super().get_form(request, obj, **kwargs)
        from BaseBillet.models import PaymentMethod
        if 'moyen_de_paiement' in form.base_fields:
            from unfold.widgets import UnfoldAdminSelectWidget
            form.base_fields['moyen_de_paiement'].widget = UnfoldAdminSelectWidget(
                choices=[('', '---')] + list(PaymentMethod.choices),
            )
        return form

    def has_add_permission(self, request):
        return TenantAdminPermissionWithRequest(request)

    def has_change_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_delete_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)

    def has_view_permission(self, request, obj=None):
        return TenantAdminPermissionWithRequest(request)
